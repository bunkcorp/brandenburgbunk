from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
OUT_PATH = ROOT / "ACTEX_PA_5.3_mechanics.xlsx"


def resolve_wage_path() -> Path:
    candidates = [
        ROOT / "Wage.csv",
        ROOT.parents[2] / "Code_Archive" / "ALL CSV FILES - 2nd Edition" / "Wage.csv",
        ROOT.parents[4] / "ISLP" / "ISLP" / "data" / "Wage.csv",
    ]
    for path in candidates:
        if path.exists():
            return path
    raise FileNotFoundError("Could not find Wage.csv from known locations.")


def main() -> None:
    wage_path = resolve_wage_path()

    # Keep workbook inspectable while still representative.
    df = pd.read_csv(wage_path).head(300).copy()
    df.insert(0, "id", range(1, len(df) + 1))
    df["wage_flag"] = (df["wage"] >= 100).astype(int)
    df["maritl_collapsed"] = df["maritl"].replace(
        {"3. Widowed": "3. Other", "4. Divorced": "3. Other", "5. Separated": "3. Other"}
    )

    # Deterministic, class-stratified split (roughly 70/30).
    train_ids: set[int] = set()
    for cls in [0, 1]:
        ids = df.loc[df["wage_flag"] == cls, "id"].tolist()
        n_train = int(round(0.7 * len(ids)))
        train_ids.update(ids[:n_train])
    df["set"] = df["id"].map(lambda x: "Train" if x in train_ids else "Test")

    wb = Workbook()
    ws_readme = wb.active
    ws_readme.title = "README"
    ws_readme.append(["sheet", "purpose"])
    ws_readme.append(["Controls", "Global parameters used by all mechanics formulas."])
    ws_readme.append(["Data", "Wage subset with wage_flag and train/test split."])
    ws_readme.append(["Transform_Tree", "Chunk 4 mechanics: wage vs sqrt(wage) stump split on age."])
    ws_readme.append(["ClassTree", "Chunks 5, 8, 9, 14 mechanics: class tree, confusion matrix."])
    ws_readme.append(["ROC_Mechanics", "Chunk 15 mechanics: threshold sweep and ROC points from class probabilities."])
    ws_readme.append(["RF_Mechanics", "Chunks 19-23 mechanics: per-tree class votes and majority prediction."])
    ws_readme.append(["Boosting_Lite", "Chunk 26-27 style additive stage logic (formula-only approximation)."])
    ws_readme.append(["CP_Pruning_Mechanics", "cp, xerror, and pruning pick mechanics tied to tree variants."])
    ws_readme.append(["RF_CV_Tuning", "Repeated CV + downsampling + mtry/ntree tuning grid mechanics."])
    ws_readme.append(["RF_Importance_PDP", "Variable-importance proxies and PDP-style views."])
    ws_readme.append(["XGB_Mechanics", "xgbTree-style tuning grid and prediction mechanics."])
    ws_readme.append(["Summary", "Key metrics pulled from all mechanics tabs."])

    ws_ctrl = wb.create_sheet("Controls")
    ws_ctrl["A1"] = "wage_high_cutoff"
    ws_ctrl["B1"] = 100
    ws_ctrl["A2"] = "classification_threshold"
    ws_ctrl["B2"] = 0.5
    ws_ctrl["A3"] = "rf_visible_trees"
    ws_ctrl["B3"] = 25
    ws_ctrl["A4"] = "boosting_stages"
    ws_ctrl["B4"] = 12
    ws_ctrl["A5"] = "boosting_learning_rate"
    ws_ctrl["B5"] = 0.2
    ws_ctrl["A6"] = "note"
    ws_ctrl["B6"] = "Workbook is formula-first mechanics, not a full package replication."
    ws_ctrl["A7"] = "cp_alpha_override (optional)"
    ws_ctrl["B7"] = ""

    ws_data = wb.create_sheet("Data")
    ws_data.append(
        [
            "id",
            "year",
            "age",
            "maritl",
            "maritl_collapsed",
            "race",
            "education",
            "jobclass",
            "health",
            "health_ins",
            "wage",
            "wage_flag",
            "set",
        ]
    )
    for row in df[
        [
            "id",
            "year",
            "age",
            "maritl",
            "maritl_collapsed",
            "race",
            "education",
            "jobclass",
            "health",
            "health_ins",
            "wage",
            "wage_flag",
            "set",
        ]
    ].itertuples(index=False):
        ws_data.append(list(row))

    n = len(df)
    data_last = 1 + n

    # Chunk 4 mechanics: compare one-split tree objective before/after sqrt transform.
    ws_t = wb.create_sheet("Transform_Tree")
    ws_t.append(
        [
            "age_threshold",
            "left_n",
            "left_mean_wage",
            "right_n",
            "right_mean_wage",
            "sse_wage",
            "left_mean_sqrtwage",
            "right_mean_sqrtwage",
            "sse_sqrtwage",
            "valid_split",
            "valid_sse_wage",
            "valid_sse_sqrt",
        ]
    )
    for r in range(2, min(data_last + 1, 252)):
        ws_t[f"A{r}"] = f"=Data!C{r}"
        ws_t[f"B{r}"] = f"=COUNTIF(Data!$C$2:$C${data_last},\"<=\"&A{r})"
        ws_t[f"C{r}"] = f"=IF(B{r}=0,0,AVERAGEIF(Data!$C$2:$C${data_last},\"<=\"&A{r},Data!$K$2:$K${data_last}))"
        ws_t[f"D{r}"] = f"=COUNTIF(Data!$C$2:$C${data_last},\">\"&A{r})"
        ws_t[f"E{r}"] = f"=IF(D{r}=0,0,AVERAGEIF(Data!$C$2:$C${data_last},\">\"&A{r},Data!$K$2:$K${data_last}))"
        ws_t[f"F{r}"] = (
            f"=SUMPRODUCT(--(Data!$C$2:$C${data_last}<=A{r}),(Data!$K$2:$K${data_last}-C{r})^2)"
            f"+SUMPRODUCT(--(Data!$C$2:$C${data_last}>A{r}),(Data!$K$2:$K${data_last}-E{r})^2)"
        )
        ws_t[f"G{r}"] = f"=IF(B{r}=0,0,SUMPRODUCT(--(Data!$C$2:$C${data_last}<=A{r}),SQRT(Data!$K$2:$K${data_last}))/B{r})"
        ws_t[f"H{r}"] = f"=IF(D{r}=0,0,SUMPRODUCT(--(Data!$C$2:$C${data_last}>A{r}),SQRT(Data!$K$2:$K${data_last}))/D{r})"
        ws_t[f"I{r}"] = (
            f"=SUMPRODUCT(--(Data!$C$2:$C${data_last}<=A{r}),(SQRT(Data!$K$2:$K${data_last})-G{r})^2)"
            f"+SUMPRODUCT(--(Data!$C$2:$C${data_last}>A{r}),(SQRT(Data!$K$2:$K${data_last})-H{r})^2)"
        )
        ws_t[f"J{r}"] = f"=IF(AND(B{r}>0,D{r}>0),1,0)"
        ws_t[f"K{r}"] = f"=IF(J{r}=1,F{r},1E+99)"
        ws_t[f"L{r}"] = f"=IF(J{r}=1,I{r},1E+99)"

    ws_t["N1"] = "best_threshold_wage"
    ws_t["N2"] = "=INDEX(A2:A251,MATCH(MIN(K2:K251),K2:K251,0))"
    ws_t["N3"] = "best_threshold_sqrt_wage"
    ws_t["N4"] = "=INDEX(A2:A251,MATCH(MIN(L2:L251),L2:L251,0))"
    ws_t["N5"] = "abs_difference"
    ws_t["N6"] = "=ABS(N2-N4)"

    # Main classification tree mechanics using age split and Gini.
    ws_c = wb.create_sheet("ClassTree")
    ws_c.append(
        [
            "age_threshold",
            "left_train_n",
            "left_train_pos",
            "right_train_n",
            "right_train_pos",
            "left_gini",
            "right_gini",
            "weighted_gini",
            "valid_split",
            "valid_gini",
        ]
    )
    for r in range(2, min(data_last + 1, 252)):
        ws_c[f"A{r}"] = f"=Data!C{r}"
        ws_c[f"B{r}"] = (
            f"=COUNTIFS(Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\"<=\"&A{r})"
        )
        ws_c[f"C{r}"] = (
            f"=SUMIFS(Data!$L$2:$L${data_last},Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\"<=\"&A{r})"
        )
        ws_c[f"D{r}"] = (
            f"=COUNTIFS(Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\">\"&A{r})"
        )
        ws_c[f"E{r}"] = (
            f"=SUMIFS(Data!$L$2:$L${data_last},Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\">\"&A{r})"
        )
        ws_c[f"F{r}"] = f"=IF(B{r}=0,0,1-(C{r}/B{r})^2-(1-C{r}/B{r})^2)"
        ws_c[f"G{r}"] = f"=IF(D{r}=0,0,1-(E{r}/D{r})^2-(1-E{r}/D{r})^2)"
        ws_c[f"H{r}"] = f"=IF(B{r}+D{r}=0,1,(B{r}*F{r}+D{r}*G{r})/(B{r}+D{r}))"
        ws_c[f"I{r}"] = f"=IF(AND(B{r}>0,D{r}>0),1,0)"
        ws_c[f"J{r}"] = f"=IF(I{r}=1,H{r},1E+99)"

    ws_c["L1"] = "best_threshold_age"
    ws_c["L2"] = "=INDEX(A2:A251,MATCH(MIN(J2:J251),J2:J251,0))"
    ws_c["L3"] = "left_prob_high"
    ws_c["L4"] = f"=IFERROR(AVERAGEIFS(Data!$L$2:$L${data_last},Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\"<=\"&$L$2),0)"
    ws_c["L5"] = "right_prob_high"
    ws_c["L6"] = f"=IFERROR(AVERAGEIFS(Data!$L$2:$L${data_last},Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\">\"&$L$2),0)"
    ws_c["L7"] = "left_class_pred"
    ws_c["L8"] = "=IF(L4>=Controls!$B$2,1,0)"
    ws_c["L9"] = "right_class_pred"
    ws_c["L10"] = "=IF(L6>=Controls!$B$2,1,0)"

    ws_c["N1"] = "test_id"
    ws_c["O1"] = "actual"
    ws_c["P1"] = "pred_prob"
    ws_c["Q1"] = "pred_class"
    ws_c["R1"] = "correct"
    test_rows = [i for i in range(2, data_last + 1) if ws_data[f"M{i}"].value == "Test"]
    for idx, drow in enumerate(test_rows, start=2):
        ws_c[f"N{idx}"] = f"=Data!A{drow}"
        ws_c[f"O{idx}"] = f"=Data!L{drow}"
        ws_c[f"P{idx}"] = f"=IF(Data!C{drow}<=$L$2,$L$4,$L$6)"
        ws_c[f"Q{idx}"] = f"=IF(P{idx}>=Controls!$B$2,1,0)"
        ws_c[f"R{idx}"] = f"=--(Q{idx}=O{idx})"
    last_pred = 1 + len(test_rows)
    class_summary_row = last_pred + 3
    ws_c[f"N{class_summary_row}"] = "accuracy"
    ws_c[f"O{class_summary_row}"] = f"=AVERAGE(R2:R{last_pred})"
    ws_c[f"N{class_summary_row+1}"] = "TP"
    ws_c[f"O{class_summary_row+1}"] = f"=COUNTIFS(O2:O{last_pred},1,Q2:Q{last_pred},1)"
    ws_c[f"N{class_summary_row+2}"] = "TN"
    ws_c[f"O{class_summary_row+2}"] = f"=COUNTIFS(O2:O{last_pred},0,Q2:Q{last_pred},0)"
    ws_c[f"N{class_summary_row+3}"] = "FP"
    ws_c[f"O{class_summary_row+3}"] = f"=COUNTIFS(O2:O{last_pred},0,Q2:Q{last_pred},1)"
    ws_c[f"N{class_summary_row+4}"] = "FN"
    ws_c[f"O{class_summary_row+4}"] = f"=COUNTIFS(O2:O{last_pred},1,Q2:Q{last_pred},0)"

    ws_roc = wb.create_sheet("ROC_Mechanics")
    ws_roc.append(["threshold", "TP", "FP", "TN", "FN", "TPR", "FPR"])
    for i in range(21):
        r = 2 + i
        thr = i / 20
        ws_roc[f"A{r}"] = thr
        ws_roc[f"B{r}"] = f"=COUNTIFS(ClassTree!$O$2:$O${last_pred},1,ClassTree!$P$2:$P${last_pred},\">=\"&A{r})"
        ws_roc[f"C{r}"] = f"=COUNTIFS(ClassTree!$O$2:$O${last_pred},0,ClassTree!$P$2:$P${last_pred},\">=\"&A{r})"
        ws_roc[f"D{r}"] = f"=COUNTIFS(ClassTree!$O$2:$O${last_pred},0,ClassTree!$P$2:$P${last_pred},\"<\"&A{r})"
        ws_roc[f"E{r}"] = f"=COUNTIFS(ClassTree!$O$2:$O${last_pred},1,ClassTree!$P$2:$P${last_pred},\"<\"&A{r})"
        ws_roc[f"F{r}"] = f"=IF(B{r}+E{r}=0,0,B{r}/(B{r}+E{r}))"
        ws_roc[f"G{r}"] = f"=IF(C{r}+D{r}=0,0,C{r}/(C{r}+D{r}))"
    ws_roc["I1"] = "approx_auc"
    ws_roc["I2"] = "=SUMPRODUCT((G2:G21-G3:G22),(F2:F21+F3:F22)/2)"
    roc_chart = ScatterChart()
    roc_chart.title = "ROC Curve (ClassTree)"
    roc_chart.x_axis.title = "False Positive Rate"
    roc_chart.y_axis.title = "True Positive Rate"
    roc_series = Series(
        Reference(ws_roc, min_col=6, min_row=2, max_row=22),
        Reference(ws_roc, min_col=7, min_row=2, max_row=22),
        title="ClassTree ROC",
    )
    roc_series.graphicalProperties.line.width = 20000
    roc_chart.series.append(roc_series)
    ws_roc.add_chart(roc_chart, "K2")

    ws_rf = wb.create_sheet("RF_Mechanics")
    ws_rf["A1"] = "tree"
    ws_rf["B1"] = "feature_used"
    ws_rf["C1"] = "threshold"
    ws_rf["D1"] = "left_prob_high"
    ws_rf["E1"] = "right_prob_high"
    ws_rf["F1"] = "left_class"
    ws_rf["G1"] = "right_class"
    n_trees = 25
    for t in range(1, n_trees + 1):
        r = 1 + t
        ws_rf[f"A{r}"] = t
        ws_rf[f"B{r}"] = f"=IF(MOD({t},2)=1,\"age\",\"year\")"
        ws_rf[f"C{r}"] = (
            f"=IF(B{r}=\"age\",INDEX(Data!$C$2:$C${data_last},1+MOD(INT(ABS(SIN(({t}*41+ROW()))*1000000)),{n})),"
            f"INDEX(Data!$B$2:$B${data_last},1+MOD(INT(ABS(SIN(({t}*53+ROW()))*1000000)),{n})))"
        )
        ws_rf[f"D{r}"] = (
            f"=IF(B{r}=\"age\","
            f"IFERROR(AVERAGEIFS(Data!$L$2:$L${data_last},Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\"<=\"&C{r}),0),"
            f"IFERROR(AVERAGEIFS(Data!$L$2:$L${data_last},Data!$M$2:$M${data_last},\"Train\",Data!$B$2:$B${data_last},\"<=\"&C{r}),0))"
        )
        ws_rf[f"E{r}"] = (
            f"=IF(B{r}=\"age\","
            f"IFERROR(AVERAGEIFS(Data!$L$2:$L${data_last},Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\">\"&C{r}),0),"
            f"IFERROR(AVERAGEIFS(Data!$L$2:$L${data_last},Data!$M$2:$M${data_last},\"Train\",Data!$B$2:$B${data_last},\">\"&C{r}),0))"
        )
        ws_rf[f"F{r}"] = f"=IF(D{r}>=Controls!$B$2,1,0)"
        ws_rf[f"G{r}"] = f"=IF(E{r}>=Controls!$B$2,1,0)"

    ws_rf["I1"] = "test_id"
    ws_rf["J1"] = "actual"
    for t in range(1, n_trees + 1):
        ws_rf.cell(row=1, column=10 + t, value=f"tree_{t:02d}_pred")
    vote_prob_col = 11 + n_trees
    maj_col = 12 + n_trees
    correct_col = 13 + n_trees
    ws_rf.cell(row=1, column=vote_prob_col, value="vote_prob_high")
    ws_rf.cell(row=1, column=maj_col, value="pred_class")
    ws_rf.cell(row=1, column=correct_col, value="correct")

    for idx, drow in enumerate(test_rows, start=2):
        ws_rf[f"I{idx}"] = f"=Data!A{drow}"
        ws_rf[f"J{idx}"] = f"=Data!L{drow}"
        for t in range(1, n_trees + 1):
            col = 10 + t
            cell = ws_rf.cell(row=idx, column=col).coordinate
            ws_rf[cell] = (
                f"=IF($B${1+t}=\"age\",IF(Data!C{drow}<=$C${1+t},$F${1+t},$G${1+t}),"
                f"IF(Data!B{drow}<=$C${1+t},$F${1+t},$G${1+t}))"
            )
        first_pred = ws_rf.cell(row=idx, column=11).coordinate
        last_pred_cell = ws_rf.cell(row=idx, column=10 + n_trees).coordinate
        vote_cell = ws_rf.cell(row=idx, column=vote_prob_col).coordinate
        pred_cell = ws_rf.cell(row=idx, column=maj_col).coordinate
        corr_cell = ws_rf.cell(row=idx, column=correct_col).coordinate
        ws_rf[vote_cell] = f"=AVERAGE({first_pred}:{last_pred_cell})"
        ws_rf[pred_cell] = f"=IF({vote_cell}>=Controls!$B$2,1,0)"
        ws_rf[corr_cell] = f"=--({pred_cell}=J{idx})"
    rf_summary_row = last_pred + 3
    ws_rf[f"I{rf_summary_row}"] = "rf_accuracy"
    ws_rf[f"J{rf_summary_row}"] = f"=AVERAGE({get_column_letter(correct_col)}2:{get_column_letter(correct_col)}{last_pred})"
    ws_rf[f"I{rf_summary_row+1}"] = "rf_tp"
    ws_rf[f"J{rf_summary_row+1}"] = f"=COUNTIFS(J2:J{last_pred},1,{get_column_letter(maj_col)}2:{get_column_letter(maj_col)}{last_pred},1)"
    ws_rf[f"I{rf_summary_row+2}"] = "rf_tn"
    ws_rf[f"J{rf_summary_row+2}"] = f"=COUNTIFS(J2:J{last_pred},0,{get_column_letter(maj_col)}2:{get_column_letter(maj_col)}{last_pred},0)"
    ws_rf[f"I{rf_summary_row+3}"] = "rf_fp"
    ws_rf[f"J{rf_summary_row+3}"] = f"=COUNTIFS(J2:J{last_pred},0,{get_column_letter(maj_col)}2:{get_column_letter(maj_col)}{last_pred},1)"
    ws_rf[f"I{rf_summary_row+4}"] = "rf_fn"
    ws_rf[f"J{rf_summary_row+4}"] = f"=COUNTIFS(J2:J{last_pred},1,{get_column_letter(maj_col)}2:{get_column_letter(maj_col)}{last_pred},0)"

    # Lightweight formula-only boosting proxy with age stumps.
    ws_boost = wb.create_sheet("Boosting_Lite")
    ws_boost["A1"] = "stage"
    ws_boost["B1"] = "threshold_age"
    ws_boost["C1"] = "left_mean_residual"
    ws_boost["D1"] = "right_mean_residual"
    ws_boost["E1"] = "eta"
    ws_boost["F1"] = "note"
    ws_boost["F2"] = "Residuals use y - p_prev on train rows."
    ws_boost["E2"] = "=Controls!$B$5"
    ws_boost["A2"] = 1
    ws_boost["B2"] = "=ClassTree!$L$2"
    ws_boost["C2"] = f"=AVERAGEIFS(Data!$L$2:$L${data_last},Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\"<=\"&B2)-ClassTree!$L$4"
    ws_boost["D2"] = f"=AVERAGEIFS(Data!$L$2:$L${data_last},Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\">\"&B2)-ClassTree!$L$6"
    for s in range(3, 14):
        ws_boost[f"A{s}"] = s - 1
        ws_boost[f"B{s}"] = f"=INDEX(ClassTree!A2:A251,1+MOD({s}*7,200))"
        ws_boost[f"C{s}"] = f"=C{s-1}*0.85"
        ws_boost[f"D{s}"] = f"=D{s-1}*0.85"
        ws_boost[f"E{s}"] = "=Controls!$B$5"

    ws_boost["H1"] = "test_id"
    ws_boost["I1"] = "actual"
    ws_boost["J1"] = "base_prob"
    for s in range(1, 13):
        ws_boost.cell(row=1, column=10 + s, value=f"p_after_s{s}")
    final_col = 10 + 12
    pred_col = final_col + 1
    corr_col = final_col + 2
    ws_boost.cell(row=1, column=pred_col, value="pred_class")
    ws_boost.cell(row=1, column=corr_col, value="correct")
    for idx, drow in enumerate(test_rows, start=2):
        ws_boost[f"H{idx}"] = f"=Data!A{drow}"
        ws_boost[f"I{idx}"] = f"=Data!L{drow}"
        ws_boost[f"J{idx}"] = f"=IF(Data!C{drow}<=ClassTree!$L$2,ClassTree!$L$4,ClassTree!$L$6)"
        prev_col_letter = "J"
        for s in range(1, 13):
            cur_col_letter = get_column_letter(10 + s)
            ws_boost[f"{cur_col_letter}{idx}"] = (
                f"=MIN(0.999,MAX(0.001,{prev_col_letter}{idx}+"
                f"INDEX($E$2:$E$13,{s})*IF(Data!C{drow}<=INDEX($B$2:$B$13,{s}),"
                f"INDEX($C$2:$C$13,{s}),INDEX($D$2:$D$13,{s}))))"
            )
            prev_col_letter = cur_col_letter
        pred_letter = get_column_letter(pred_col)
        corr_letter = get_column_letter(corr_col)
        final_letter = get_column_letter(final_col)
        ws_boost[f"{pred_letter}{idx}"] = f"=IF({final_letter}{idx}>=Controls!$B$2,1,0)"
        ws_boost[f"{corr_letter}{idx}"] = f"=--({pred_letter}{idx}=I{idx})"
    boost_summary_row = last_pred + 3
    ws_boost[f"H{boost_summary_row}"] = "boost_lite_accuracy"
    ws_boost[f"I{boost_summary_row}"] = f"=AVERAGE({get_column_letter(corr_col)}2:{get_column_letter(corr_col)}{last_pred})"

    # cp + xerror + pruning mechanics.
    ws_cp = wb.create_sheet("CP_Pruning_Mechanics")
    ws_cp.append(["model", "nsplit", "rel_error", "xerror", "xstd", "cp"])
    ws_cp["A2"] = "root"
    ws_cp["B2"] = 0
    ws_cp["C2"] = 1
    ws_cp["D2"] = f"=1-COUNTIFS(Data!$M$2:$M${data_last},\"Test\",Data!$L$2:$L${data_last},IF(AVERAGEIFS(Data!$L$2:$L${data_last},Data!$M$2:$M${data_last},\"Train\")>=0.5,1,0))/COUNTIF(Data!$M$2:$M${data_last},\"Test\")"
    ws_cp["E2"] = "=SQRT(D2*(1-D2)/COUNTIF(Data!$M$2:$M$301,\"Test\"))"
    ws_cp["F2"] = "=1"
    ws_cp["A3"] = "stump"
    ws_cp["B3"] = 1
    ws_cp["C3"] = "=1-ClassTree!O{0}".format(class_summary_row)
    ws_cp["D3"] = "=1-ClassTree!O{0}".format(class_summary_row)
    ws_cp["E3"] = "=SQRT(D3*(1-D3)/COUNTIF(Data!$M$2:$M$301,\"Test\"))"
    ws_cp["A4"] = "depth2_age_year"
    ws_cp["B4"] = 3
    ws_cp["C4"] = "=MAX(0.0001,C3*0.9)"
    ws_cp["L1"] = "year_threshold_candidate"
    ws_cp["M1"] = "left_branch_gini"
    ws_cp["N1"] = "right_branch_gini"
    ws_cp["O1"] = "weighted_gini"
    ws_cp["P1"] = "valid_weighted_gini"
    for r in range(2, 152):
        drow = r
        ws_cp[f"L{r}"] = f"=Data!B{drow}"
        ws_cp[f"M{r}"] = (
            f"=IF(COUNTIFS(Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\"<=\"&ClassTree!$L$2,Data!$B$2:$B${data_last},\"<=\"&L{r})=0,0,"
            f"1-(SUMIFS(Data!$L$2:$L${data_last},Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\"<=\"&ClassTree!$L$2,Data!$B$2:$B${data_last},\"<=\"&L{r})/"
            f"COUNTIFS(Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\"<=\"&ClassTree!$L$2,Data!$B$2:$B${data_last},\"<=\"&L{r}))^2-"
            f"(1-SUMIFS(Data!$L$2:$L${data_last},Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\"<=\"&ClassTree!$L$2,Data!$B$2:$B${data_last},\"<=\"&L{r})/"
            f"COUNTIFS(Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\"<=\"&ClassTree!$L$2,Data!$B$2:$B${data_last},\"<=\"&L{r}))^2)"
        )
        ws_cp[f"N{r}"] = (
            f"=IF(COUNTIFS(Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\">\"&ClassTree!$L$2,Data!$B$2:$B${data_last},\">\"&L{r})=0,0,"
            f"1-(SUMIFS(Data!$L$2:$L${data_last},Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\">\"&ClassTree!$L$2,Data!$B$2:$B${data_last},\">\"&L{r})/"
            f"COUNTIFS(Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\">\"&ClassTree!$L$2,Data!$B$2:$B${data_last},\">\"&L{r}))^2-"
            f"(1-SUMIFS(Data!$L$2:$L${data_last},Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\">\"&ClassTree!$L$2,Data!$B$2:$B${data_last},\">\"&L{r})/"
            f"COUNTIFS(Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\">\"&ClassTree!$L$2,Data!$B$2:$B${data_last},\">\"&L{r}))^2)"
        )
        ws_cp[f"O{r}"] = (
            f"=((COUNTIFS(Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\"<=\"&ClassTree!$L$2)*M{r})+"
            f"(COUNTIFS(Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\">\"&ClassTree!$L$2)*N{r}))/"
            f"COUNTIF(Data!$M$2:$M${data_last},\"Train\")"
        )
        ws_cp[f"P{r}"] = f"=IF(AND(COUNTIFS(Data!$M$2:$M${data_last},\"Train\",Data!$B$2:$B${data_last},\"<=\"&L{r})>0,COUNTIFS(Data!$M$2:$M${data_last},\"Train\",Data!$B$2:$B${data_last},\">\"&L{r})>0),O{r},1E+99)"
    ws_cp["R1"] = "best_year_split"
    ws_cp["R2"] = "=INDEX(L2:L151,MATCH(MIN(P2:P151),P2:P151,0))"
    ws_cp["R3"] = "depth2_test_accuracy"
    ws_cp["R4"] = (
        f"=AVERAGE(--(IF(Data!$M$2:$M${data_last}=\"Test\","
        f"IF(Data!$C$2:$C${data_last}<=ClassTree!$L$2,"
        f"IF(Data!$B$2:$B${data_last}<=R2,"
        f"IF(IFERROR(AVERAGEIFS(Data!$L$2:$L${data_last},Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\"<=\"&ClassTree!$L$2,Data!$B$2:$B${data_last},\"<=\"&R2),0)>=0.5,1,0),"
        f"IF(IFERROR(AVERAGEIFS(Data!$L$2:$L${data_last},Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\"<=\"&ClassTree!$L$2,Data!$B$2:$B${data_last},\">\"&R2),0)>=0.5,1,0)),"
        f"IF(Data!$B$2:$B${data_last}<=R2,"
        f"IF(IFERROR(AVERAGEIFS(Data!$L$2:$L${data_last},Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\">\"&ClassTree!$L$2,Data!$B$2:$B${data_last},\"<=\"&R2),0)>=0.5,1,0),"
        f"IF(IFERROR(AVERAGEIFS(Data!$L$2:$L${data_last},Data!$M$2:$M${data_last},\"Train\",Data!$C$2:$C${data_last},\">\"&ClassTree!$L$2,Data!$B$2:$B${data_last},\">\"&R2),0)>=0.5,1,0))),"
        f"\"\"))=IF(Data!$M$2:$M${data_last}=\"Test\",Data!$L$2:$L${data_last},\"\")))"
    )
    ws_cp["D4"] = "=1-R4"
    ws_cp["E4"] = "=SQRT(D4*(1-D4)/COUNTIF(Data!$M$2:$M$301,\"Test\"))"
    ws_cp["F3"] = "=MAX(0.0001,D2-D3)"
    ws_cp["F4"] = "=MAX(0.0001,(D3-D4)/2)"
    ws_cp["H1"] = "cp_min_xerror"
    ws_cp["H2"] = "=INDEX(F2:F4,MATCH(MIN(D2:D4),D2:D4,0))"
    ws_cp["H3"] = "min_xerror_model"
    ws_cp["H4"] = "=INDEX(A2:A4,MATCH(MIN(D2:D4),D2:D4,0))"
    ws_cp["H5"] = "pruned_model_by_cp"
    ws_cp["H6"] = "=IF(Controls!$B$7=\"\",H4,IF(Controls!$B$7>=F2,\"root\",IF(Controls!$B$7>=F3,\"stump\",\"depth2_age_year\")))"
    cp_chart = LineChart()
    cp_chart.title = "CP Table: xerror by subtree"
    cp_chart.y_axis.title = "xerror"
    cp_chart.x_axis.title = "nsplit"
    cp_chart.add_data(Reference(ws_cp, min_col=4, min_row=2, max_row=4), titles_from_data=False)
    cp_chart.set_categories(Reference(ws_cp, min_col=2, min_row=2, max_row=4))
    ws_cp.add_chart(cp_chart, "A8")

    # Repeated CV + tuning mechanics.
    ws_cv = wb.create_sheet("RF_CV_Tuning")
    ws_cv["A1"] = "train_id"
    ws_cv["B1"] = "wage_flag"
    ws_cv["C1"] = "fold_rep1"
    ws_cv["D1"] = "fold_rep2"
    ws_cv["E1"] = "fold_rep3"
    train_rows = [i for i in range(2, data_last + 1) if ws_data[f"M{i}"].value == "Train"]
    for idx, drow in enumerate(train_rows, start=2):
        ws_cv[f"A{idx}"] = f"=Data!A{drow}"
        ws_cv[f"B{idx}"] = f"=Data!L{drow}"
        ws_cv[f"C{idx}"] = f"=1+MOD(A{idx}-1,5)"
        ws_cv[f"D{idx}"] = f"=1+MOD(A{idx}+1,5)"
        ws_cv[f"E{idx}"] = f"=1+MOD(A{idx}+3,5)"
    cv_last = 1 + len(train_rows)
    ws_cv["G1"] = "downsample_count_per_fold_rep1"
    ws_cv["H1"] = "downsample_count_per_fold_rep2"
    ws_cv["I1"] = "downsample_count_per_fold_rep3"
    for f in range(1, 6):
        r = 1 + f
        ws_cv[f"G{r}"] = f"=MIN(COUNTIFS(C2:C{cv_last},{f},B2:B{cv_last},0),COUNTIFS(C2:C{cv_last},{f},B2:B{cv_last},1))"
        ws_cv[f"H{r}"] = f"=MIN(COUNTIFS(D2:D{cv_last},{f},B2:B{cv_last},0),COUNTIFS(D2:D{cv_last},{f},B2:B{cv_last},1))"
        ws_cv[f"I{r}"] = f"=MIN(COUNTIFS(E2:E{cv_last},{f},B2:B{cv_last},0),COUNTIFS(E2:E{cv_last},{f},B2:B{cv_last},1))"
    ws_cv["K1"] = "repeat"
    ws_cv["L1"] = "fold"
    ws_cv["M1"] = "mtry"
    ws_cv["N1"] = "ntree"
    ws_cv["O1"] = "cv_score_proxy"
    row = 2
    for rep in [1, 2, 3]:
        for fold in [1, 2, 3, 4, 5]:
            for mtry in [1, 2, 3, 4, 5]:
                for ntree in [5, 20, 100]:
                    ws_cv[f"K{row}"] = rep
                    ws_cv[f"L{row}"] = fold
                    ws_cv[f"M{row}"] = mtry
                    ws_cv[f"N{row}"] = ntree
                    ws_cv[f"O{row}"] = (
                        f"=MAX(0.5,MIN(0.99,RF_Mechanics!J{rf_summary_row}"
                        f"-ABS(M{row}-2.5)*0.015"
                        f"+LOG10(N{row})*0.02"
                        f"-ABS(L{row}-3)*0.003"
                        f"-ABS(K{row}-2)*0.002))"
                    )
                    row += 1
    cv_grid_last = row - 1
    ws_cv["Q1"] = "mtry"
    ws_cv["R1"] = "mean_cv_score"
    for mtry in [1, 2, 3, 4, 5]:
        rr = 1 + mtry
        ws_cv[f"Q{rr}"] = mtry
        ws_cv[f"R{rr}"] = f"=AVERAGEIFS(O2:O{cv_grid_last},M2:M{cv_grid_last},Q{rr})"
    ws_cv["Q8"] = "best_mtry"
    ws_cv["R8"] = "=INDEX(Q2:Q6,MATCH(MAX(R2:R6),R2:R6,0))"
    cv_chart = LineChart()
    cv_chart.title = "RF Tuning Curve (mean CV score by mtry)"
    cv_chart.y_axis.title = "mean CV score"
    cv_chart.x_axis.title = "mtry"
    cv_chart.add_data(Reference(ws_cv, min_col=18, min_row=1, max_row=6), titles_from_data=True)
    cv_chart.set_categories(Reference(ws_cv, min_col=17, min_row=2, max_row=6))
    ws_cv.add_chart(cv_chart, "T2")

    # Variable importance + PDP-style sheets.
    ws_imp = wb.create_sheet("RF_Importance_PDP")
    ws_imp["A1"] = "variable"
    ws_imp["B1"] = "importance_proxy"
    ws_imp["A2"] = "age"
    ws_imp["B2"] = f"=ABS(CORREL(Data!$C$2:$C${data_last},Data!$L$2:$L${data_last}))"
    ws_imp["A3"] = "year"
    ws_imp["B3"] = f"=ABS(CORREL(Data!$B$2:$B${data_last},Data!$L$2:$L${data_last}))"
    ws_imp["A4"] = "education"
    ws_imp["B4"] = (
        f"=MAX(ABS(AVERAGEIFS(Data!$L$2:$L${data_last},Data!$G$2:$G${data_last},\"1. < HS Grad\")-AVERAGE(Data!$L$2:$L${data_last})),"
        f"ABS(AVERAGEIFS(Data!$L$2:$L${data_last},Data!$G$2:$G${data_last},\"5. Advanced Degree\")-AVERAGE(Data!$L$2:$L${data_last})))"
    )
    ws_imp["A5"] = "jobclass"
    ws_imp["B5"] = f"=ABS(AVERAGEIFS(Data!$L$2:$L${data_last},Data!$H$2:$H${data_last},\"2. Information\")-AVERAGEIFS(Data!$L$2:$L${data_last},Data!$H$2:$H${data_last},\"1. Industrial\"))"
    ws_imp["A6"] = "health_ins"
    ws_imp["B6"] = f"=ABS(AVERAGEIFS(Data!$L$2:$L${data_last},Data!$J$2:$J${data_last},\"1. Yes\")-AVERAGEIFS(Data!$L$2:$L${data_last},Data!$J$2:$J${data_last},\"2. No\"))"
    ws_imp["A7"] = "maritl_collapsed"
    ws_imp["B7"] = (
        f"=MAX(ABS(AVERAGEIFS(Data!$L$2:$L${data_last},Data!$E$2:$E${data_last},\"1. Never Married\")-AVERAGE(Data!$L$2:$L${data_last})),"
        f"ABS(AVERAGEIFS(Data!$L$2:$L${data_last},Data!$E$2:$E${data_last},\"3. Other\")-AVERAGE(Data!$L$2:$L${data_last})))"
    )
    imp_chart = BarChart()
    imp_chart.title = "RF Variable Importance (Proxy)"
    imp_chart.add_data(Reference(ws_imp, min_col=2, min_row=1, max_row=7), titles_from_data=True)
    imp_chart.set_categories(Reference(ws_imp, min_col=1, min_row=2, max_row=7))
    ws_imp.add_chart(imp_chart, "D1")

    ws_imp["A10"] = "education_level"
    ws_imp["B10"] = "partial_dependence_proxy"
    edu_levels = [
        "1. < HS Grad",
        "2. HS Grad",
        "3. Some College",
        "4. College Grad",
        "5. Advanced Degree",
    ]
    for i, level in enumerate(edu_levels, start=11):
        ws_imp[f"A{i}"] = level
        ws_imp[f"B{i}"] = f"=AVERAGEIFS(Data!$L$2:$L${data_last},Data!$M$2:$M${data_last},\"Train\",Data!$G$2:$G${data_last},A{i})"

    ws_imp["A18"] = "age_grid"
    ws_imp["B18"] = "partial_dependence_proxy"
    for i in range(19, 40):
        ws_imp[f"A{i}"] = 18 + (i - 19) * 2
        ws_imp[f"B{i}"] = f"=IF(A{i}<=ClassTree!$L$2,ClassTree!$L$4,ClassTree!$L$6)"
    pdp_chart = LineChart()
    pdp_chart.title = "PDP Proxy: age"
    pdp_chart.add_data(Reference(ws_imp, min_col=2, min_row=18, max_row=39), titles_from_data=True)
    pdp_chart.set_categories(Reference(ws_imp, min_col=1, min_row=19, max_row=39))
    ws_imp.add_chart(pdp_chart, "D18")

    # XGBoost-style tuning + prediction mechanics.
    ws_xgb = wb.create_sheet("XGB_Mechanics")
    ws_xgb["A1"] = "max_depth"
    ws_xgb["B1"] = "min_child_weight"
    ws_xgb["C1"] = "gamma"
    ws_xgb["D1"] = "nrounds"
    ws_xgb["E1"] = "eta"
    ws_xgb["F1"] = "colsample_bytree"
    ws_xgb["G1"] = "subsample"
    ws_xgb["H1"] = "cv_score_proxy"
    row = 2
    for nround in [50, 100, 150, 200, 250]:
        for eta in [0.001, 0.002, 0.01, 0.02, 0.1]:
            ws_xgb[f"A{row}"] = 7
            ws_xgb[f"B{row}"] = 1
            ws_xgb[f"C{row}"] = 0
            ws_xgb[f"D{row}"] = nround
            ws_xgb[f"E{row}"] = eta
            ws_xgb[f"F{row}"] = 0.6
            ws_xgb[f"G{row}"] = 0.6
            ws_xgb[f"H{row}"] = (
                f"=MAX(0.5,MIN(0.99,Boosting_Lite!I{boost_summary_row}"
                f"-ABS(D{row}-150)/800"
                f"-ABS(LOG10(E{row})+1.7)*0.04))"
            )
            row += 1
    xgb_last = row - 1
    ws_xgb["J1"] = "best_nrounds"
    ws_xgb["J2"] = f"=INDEX(D2:D{xgb_last},MATCH(MAX(H2:H{xgb_last}),H2:H{xgb_last},0))"
    ws_xgb["J3"] = "best_eta"
    ws_xgb["J4"] = f"=INDEX(E2:E{xgb_last},MATCH(MAX(H2:H{xgb_last}),H2:H{xgb_last},0))"
    ws_xgb["J5"] = "best_cv_score"
    ws_xgb["J6"] = f"=MAX(H2:H{xgb_last})"
    ws_xgb["A30"] = "test_id"
    ws_xgb["B30"] = "actual"
    ws_xgb["C30"] = "pred_prob_xgb_proxy"
    ws_xgb["D30"] = "pred_class"
    ws_xgb["E30"] = "correct"
    for idx, drow in enumerate(test_rows, start=31):
        ws_xgb[f"A{idx}"] = f"=Data!A{drow}"
        ws_xgb[f"B{idx}"] = f"=Data!L{drow}"
        ws_xgb[f"C{idx}"] = (
            f"=MIN(0.999,MAX(0.001,ClassTree!P{idx-29}"
            f"+0.15*LOG10($J$2)"
            f"+0.08*($J$4-0.02)"
            f"+0.03*IF(Data!C{drow}>ClassTree!$L$2,1,-1)))"
        )
        ws_xgb[f"D{idx}"] = f"=IF(C{idx}>=Controls!$B$2,1,0)"
        ws_xgb[f"E{idx}"] = f"=--(D{idx}=B{idx})"
    xgb_pred_last = 30 + len(test_rows)
    xgb_summary_row = xgb_pred_last + 3
    ws_xgb[f"A{xgb_summary_row}"] = "xgb_accuracy"
    ws_xgb[f"B{xgb_summary_row}"] = f"=AVERAGE(E31:E{xgb_pred_last})"
    ws_xgb[f"A{xgb_summary_row+1}"] = "TP"
    ws_xgb[f"B{xgb_summary_row+1}"] = f"=COUNTIFS(B31:B{xgb_pred_last},1,D31:D{xgb_pred_last},1)"
    ws_xgb[f"A{xgb_summary_row+2}"] = "TN"
    ws_xgb[f"B{xgb_summary_row+2}"] = f"=COUNTIFS(B31:B{xgb_pred_last},0,D31:D{xgb_pred_last},0)"
    ws_xgb[f"A{xgb_summary_row+3}"] = "FP"
    ws_xgb[f"B{xgb_summary_row+3}"] = f"=COUNTIFS(B31:B{xgb_pred_last},0,D31:D{xgb_pred_last},1)"
    ws_xgb[f"A{xgb_summary_row+4}"] = "FN"
    ws_xgb[f"B{xgb_summary_row+4}"] = f"=COUNTIFS(B31:B{xgb_pred_last},1,D31:D{xgb_pred_last},0)"
    ws_xgb["G30"] = "threshold"
    ws_xgb["H30"] = "TPR"
    ws_xgb["I30"] = "FPR"
    for i in range(21):
        rr = 31 + i
        ws_xgb[f"G{rr}"] = i / 20
        ws_xgb[f"H{rr}"] = f"=IF(COUNTIF(B31:B{xgb_pred_last},1)=0,0,COUNTIFS(B31:B{xgb_pred_last},1,C31:C{xgb_pred_last},\">=\"&G{rr})/COUNTIF(B31:B{xgb_pred_last},1))"
        ws_xgb[f"I{rr}"] = f"=IF(COUNTIF(B31:B{xgb_pred_last},0)=0,0,COUNTIFS(B31:B{xgb_pred_last},0,C31:C{xgb_pred_last},\">=\"&G{rr})/COUNTIF(B31:B{xgb_pred_last},0))"
    ws_xgb["K30"] = "auc_proxy"
    ws_xgb["K31"] = "=SUMPRODUCT((I31:I50-I32:I51),(H31:H50+H32:H51)/2)"
    xgb_curve = LineChart()
    xgb_curve.title = "XGB Tuning Curve (score vs nrounds)"
    xgb_curve.y_axis.title = "cv score proxy"
    xgb_curve.x_axis.title = "nrounds"
    xgb_curve.add_data(Reference(ws_xgb, min_col=8, min_row=1, max_row=26), titles_from_data=True)
    xgb_curve.set_categories(Reference(ws_xgb, min_col=4, min_row=2, max_row=26))
    ws_xgb.add_chart(xgb_curve, "M2")
    xgb_roc = ScatterChart()
    xgb_roc.title = "ROC Curve (XGB proxy)"
    xgb_roc.x_axis.title = "FPR"
    xgb_roc.y_axis.title = "TPR"
    xgb_roc.series.append(
        Series(
            Reference(ws_xgb, min_col=8, min_row=31, max_row=51),
            Reference(ws_xgb, min_col=9, min_row=31, max_row=51),
            title="XGB ROC",
        )
    )
    ws_xgb.add_chart(xgb_roc, "M20")

    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["metric", "value"])
    ws_sum.append(["n_rows_data", n])
    ws_sum.append(["train_rows", f"=COUNTIF(Data!M2:M{data_last},\"Train\")"])
    ws_sum.append(["test_rows", f"=COUNTIF(Data!M2:M{data_last},\"Test\")"])
    ws_sum.append(["class_tree_accuracy", f"=ClassTree!O{class_summary_row}"])
    ws_sum.append(["class_tree_auc_approx", "=ROC_Mechanics!I2"])
    ws_sum.append(["cp_min_xerror", "=CP_Pruning_Mechanics!H2"])
    ws_sum.append(["cp_pruned_model", "=CP_Pruning_Mechanics!H6"])
    ws_sum.append(["rf_best_mtry_cv", "=RF_CV_Tuning!R8"])
    ws_sum.append(["rf_vote_accuracy", f"=RF_Mechanics!J{rf_summary_row}"])
    ws_sum.append(["xgb_best_cv_score_proxy", "=XGB_Mechanics!J6"])
    ws_sum.append(["xgb_accuracy_proxy", f"=XGB_Mechanics!B{xgb_summary_row}"])
    ws_sum.append(["xgb_auc_proxy", "=XGB_Mechanics!K31"])
    ws_sum.append(["boost_lite_accuracy", f"=Boosting_Lite!I{boost_summary_row}"])
    ws_sum.append(["best_age_split_wage", "=Transform_Tree!N2"])
    ws_sum.append(["best_age_split_sqrt_wage", "=Transform_Tree!N4"])

    for ws in wb.worksheets:
        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=30):
            for cell in col:
                if cell.value is not None:
                    ws.column_dimensions[cell.column_letter].width = max(
                        ws.column_dimensions[cell.column_letter].width or 0,
                        min(max(len(str(cell.value)) + 2, 12), 42),
                    )

    wb.save(OUT_PATH)
    print(f"Wrote: {OUT_PATH}")


if __name__ == "__main__":
    main()
