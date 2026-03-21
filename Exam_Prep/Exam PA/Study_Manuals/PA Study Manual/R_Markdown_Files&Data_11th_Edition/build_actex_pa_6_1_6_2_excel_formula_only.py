from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

try:
    from scipy.cluster.hierarchy import dendrogram, linkage
except ImportError:  # pragma: no cover
    dendrogram = None
    linkage = None


ROOT = Path(__file__).resolve().parent
OUT_PATH = ROOT / "ACTEX_PA_6.1_6.2_mechanics.xlsx"


def resolve_credit_path() -> Path:
    candidates = [
        ROOT / "Credit.csv",
        ROOT / "credit_full.csv",
        ROOT.parents[4] / "ISLP" / "ISLP" / "data" / "Credit.csv",
        ROOT.parents[2] / "Code_Archive" / "ALL CSV FILES - 2nd Edition" / "Credit.csv",
    ]
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError("Could not find Credit.csv.")


def load_usarrests() -> pd.DataFrame:
    urls = [
        "https://raw.githubusercontent.com/selva86/datasets/master/USArrests.csv",
        "https://raw.githubusercontent.com/vincentarelbundock/Rdatasets/master/csv/datasets/USArrests.csv",
    ]
    df = None
    for url in urls:
        try:
            df = pd.read_csv(url)
            break
        except Exception:
            continue
    if df is None:
        raise RuntimeError("Unable to download USArrests.")
    if "state" not in df.columns:
        if "State" in df.columns:
            df = df.rename(columns={"State": "state"})
        elif "rownames" in df.columns:
            df = df.rename(columns={"rownames": "state"})
        else:
            df = df.rename(columns={df.columns[0]: "state"})
    return df[["state", "Murder", "Assault", "UrbanPop", "Rape"]].copy()


def main() -> None:
    usa = load_usarrests()
    z_np = (
        (usa[["Murder", "Assault", "UrbanPop", "Rape"]] - usa[["Murder", "Assault", "UrbanPop", "Rape"]].mean())
        / usa[["Murder", "Assault", "UrbanPop", "Rape"]].std(ddof=1)
    ).to_numpy()
    credit = pd.read_csv(resolve_credit_path())
    if {"Limit", "Rating"}.issubset(credit.columns):
        credit = credit[["Limit", "Rating"]].dropna().head(400).copy()
    else:
        credit = credit.iloc[:, :2].dropna().head(400).copy()
        credit.columns = ["Limit", "Rating"]

    wb = Workbook()
    ws_readme = wb.active
    ws_readme.title = "README"
    ws_readme.append(["sheet", "purpose"])
    ws_readme.append(["Data_USArrests", "Raw USArrests data only."])
    ws_readme.append(["Scale_PCA_Mechanics", "Scaling, correlation matrix, PC1 power iteration, scores/PVE via formulas."])
    ws_readme.append(["Plot_Parity_6_1", "6.1 visual parity charts: histograms, PVE/CumPVE, biplot-style views."])
    ws_readme.append(["KMeans_6_2", "Formula-based k=3 clustering iteration and elbow proxy."])
    ws_readme.append(["Cluster_Summary", "Formula-based grouped means by final k=3 cluster."])
    ws_readme.append(["Distance_HC_Mechanics", "Pairwise distances, closest pair, and linkage metric mechanics."])
    ws_readme.append(["Plot_Parity_6_2", "6.2 visual parity charts: cluster PC plot + linkage profiles."])
    ws_readme.append(["Credit_PCA_Mechanics", "Limit/Rating PCA mechanics with formulas."])
    ws_readme.append(["Summary", "Key checks and metrics from all tabs."])

    ws_data = wb.create_sheet("Data_USArrests")
    ws_data.append(["id", "state", "Murder", "Assault", "UrbanPop", "Rape"])
    for i, row in enumerate(usa.itertuples(index=False), start=1):
        ws_data.append([i, row.state, float(row.Murder), float(row.Assault), float(row.UrbanPop), float(row.Rape)])

    # ---------------------------
    # PCA mechanics (formula-first)
    # ---------------------------
    ws_pca = wb.create_sheet("Scale_PCA_Mechanics")
    ws_pca.append(
        [
            "id",
            "Murder",
            "Assault",
            "UrbanPop",
            "Rape",
            "",
            "z_Murder",
            "z_Assault",
            "z_UrbanPop",
            "z_Rape",
            "",
            "PC1_score",
        ]
    )
    for r in range(2, 52):
        ws_pca[f"A{r}"] = f"=Data_USArrests!A{r}"
        ws_pca[f"B{r}"] = f"=Data_USArrests!C{r}"
        ws_pca[f"C{r}"] = f"=Data_USArrests!D{r}"
        ws_pca[f"D{r}"] = f"=Data_USArrests!E{r}"
        ws_pca[f"E{r}"] = f"=Data_USArrests!F{r}"

    ws_pca["A54"] = "mean"
    ws_pca["A55"] = "sd"
    for col in ["B", "C", "D", "E"]:
        ws_pca[f"{col}54"] = f"=AVERAGE({col}2:{col}51)"
        ws_pca[f"{col}55"] = f"=STDEV({col}2:{col}51)"
    for r in range(2, 52):
        ws_pca[f"G{r}"] = f"=(B{r}-$B$54)/$B$55"
        ws_pca[f"H{r}"] = f"=(C{r}-$C$54)/$C$55"
        ws_pca[f"I{r}"] = f"=(D{r}-$D$54)/$D$55"
        ws_pca[f"J{r}"] = f"=(E{r}-$E$54)/$E$55"

    ws_pca["M1"] = "Correlation Matrix R"
    ws_pca["M2"] = "Murder"
    ws_pca["N2"] = "Assault"
    ws_pca["O2"] = "UrbanPop"
    ws_pca["P2"] = "Rape"
    ws_pca["L3"] = "Murder"
    ws_pca["L4"] = "Assault"
    ws_pca["L5"] = "UrbanPop"
    ws_pca["L6"] = "Rape"
    zcols = ["G", "H", "I", "J"]
    for i, rowcol in enumerate(zcols, start=3):
        for j, colcol in enumerate(zcols, start=13):
            c = chr(64 + j)
            ws_pca[f"{c}{i}"] = f"=SUMPRODUCT({rowcol}2:{rowcol}51,{colcol}2:{colcol}51)/(COUNT({rowcol}2:{rowcol}51)-1)"

    ws_pca["L9"] = "PC1 power iteration"
    ws_pca["L10"] = "iter"
    ws_pca["M10"] = "v1"
    ws_pca["N10"] = "v2"
    ws_pca["O10"] = "v3"
    ws_pca["P10"] = "v4"
    ws_pca["Q10"] = "t1"
    ws_pca["R10"] = "t2"
    ws_pca["S10"] = "t3"
    ws_pca["T10"] = "t4"
    ws_pca["U10"] = "norm_t"
    ws_pca["L11"] = 0
    ws_pca["M11"] = 0.5
    ws_pca["N11"] = 0.5
    ws_pca["O11"] = 0.5
    ws_pca["P11"] = 0.5
    for r in range(12, 31):
        prev = r - 1
        ws_pca[f"L{r}"] = r - 11
        ws_pca[f"Q{r}"] = f"=SUMPRODUCT($M$3:$P$3,M{prev}:P{prev})"
        ws_pca[f"R{r}"] = f"=SUMPRODUCT($M$4:$P$4,M{prev}:P{prev})"
        ws_pca[f"S{r}"] = f"=SUMPRODUCT($M$5:$P$5,M{prev}:P{prev})"
        ws_pca[f"T{r}"] = f"=SUMPRODUCT($M$6:$P$6,M{prev}:P{prev})"
        ws_pca[f"U{r}"] = f"=SQRT(SUMSQ(Q{r}:T{r}))"
        ws_pca[f"M{r}"] = f"=Q{r}/U{r}"
        ws_pca[f"N{r}"] = f"=R{r}/U{r}"
        ws_pca[f"O{r}"] = f"=S{r}/U{r}"
        ws_pca[f"P{r}"] = f"=T{r}/U{r}"

    ws_pca["L33"] = "final_PC1_loading"
    ws_pca["M33"] = "=M30"
    ws_pca["N33"] = "=N30"
    ws_pca["O33"] = "=O30"
    ws_pca["P33"] = "=P30"
    ws_pca["L34"] = "lambda1"
    ws_pca["M34"] = (
        "=M33*(M3*M33+N3*N33+O3*O33+P3*P33)+"
        "N33*(M4*M33+N4*N33+O4*O33+P4*P33)+"
        "O33*(M5*M33+N5*N33+O5*O33+P5*P33)+"
        "P33*(M6*M33+N6*N33+O6*O33+P6*P33)"
    )
    ws_pca["L35"] = "PVE_PC1"
    ws_pca["M35"] = "=M34/4"

    for r in range(2, 52):
        ws_pca[f"L{r}"] = f"=SUMPRODUCT(G{r}:J{r},$M$33:$P$33)"
    ws_pca["N37"] = "Alabama_manual"
    ws_pca["O37"] = "=SUMPRODUCT(G2:J2,$M$33:$P$33)"
    ws_pca["N38"] = "Alabama_from_scores"
    ws_pca["O38"] = "=L2"
    ws_pca["N39"] = "difference"
    ws_pca["O39"] = "=O37-O38"

    # PC2 via deflation + second power iteration (formula-only).
    ws_pca["AA1"] = "Deflated matrix for PC2"
    for i in range(4):
        for j in range(4):
            row = 3 + i
            col = 27 + j  # AA..AD
            col_letter = get_column_letter(col)
            v1_i = chr(77 + i)  # M..P
            v1_j = chr(77 + j)
            r_cell = chr(77 + j) + str(3 + i)  # M3:P6
            ws_pca[f"{col_letter}{row}"] = f"={r_cell}-$M$34*${v1_i}$33*${v1_j}$33"

    ws_pca["AA10"] = "PC2 power iteration"
    ws_pca["AA11"] = 0
    ws_pca["AB10"] = "v1"
    ws_pca["AC10"] = "v2"
    ws_pca["AD10"] = "v3"
    ws_pca["AE10"] = "v4"
    ws_pca["AF10"] = "t1"
    ws_pca["AG10"] = "t2"
    ws_pca["AH10"] = "t3"
    ws_pca["AI10"] = "t4"
    ws_pca["AJ10"] = "norm_t"
    ws_pca["AB11"] = 0.5
    ws_pca["AC11"] = -0.5
    ws_pca["AD11"] = 0.5
    ws_pca["AE11"] = -0.5
    for r in range(12, 31):
        prev = r - 1
        ws_pca[f"AA{r}"] = r - 11
        ws_pca[f"AF{r}"] = f"=SUMPRODUCT($AA$3:$AD$3,AB{prev}:AE{prev})"
        ws_pca[f"AG{r}"] = f"=SUMPRODUCT($AA$4:$AD$4,AB{prev}:AE{prev})"
        ws_pca[f"AH{r}"] = f"=SUMPRODUCT($AA$5:$AD$5,AB{prev}:AE{prev})"
        ws_pca[f"AI{r}"] = f"=SUMPRODUCT($AA$6:$AD$6,AB{prev}:AE{prev})"
        ws_pca[f"AJ{r}"] = f"=SQRT(SUMSQ(AF{r}:AI{r}))"
        ws_pca[f"AB{r}"] = f"=AF{r}/AJ{r}"
        ws_pca[f"AC{r}"] = f"=AG{r}/AJ{r}"
        ws_pca[f"AD{r}"] = f"=AH{r}/AJ{r}"
        ws_pca[f"AE{r}"] = f"=AI{r}/AJ{r}"

    ws_pca["AA33"] = "final_PC2_loading"
    ws_pca["AB33"] = "=AB30"
    ws_pca["AC33"] = "=AC30"
    ws_pca["AD33"] = "=AD30"
    ws_pca["AE33"] = "=AE30"
    ws_pca["AA34"] = "lambda2"
    ws_pca["AB34"] = (
        "=AB33*(M3*AB33+N3*AC33+O3*AD33+P3*AE33)+"
        "AC33*(M4*AB33+N4*AC33+O4*AD33+P4*AE33)+"
        "AD33*(M5*AB33+N5*AC33+O5*AD33+P5*AE33)+"
        "AE33*(M6*AB33+N6*AC33+O6*AD33+P6*AE33)"
    )
    ws_pca["AA35"] = "PVE_PC2"
    ws_pca["AB35"] = "=AB34/4"
    ws_pca["Z1"] = "PC2_score"
    for r in range(2, 52):
        ws_pca[f"Z{r}"] = f"=SUMPRODUCT(G{r}:J{r},$AB$33:$AE$33)"
    ws_pca["W10"] = "PC"
    ws_pca["X10"] = "PVE_proxy"
    ws_pca["Y10"] = "CumPVE_proxy"
    ws_pca["W11"] = "PC1"
    ws_pca["X11"] = "=M35"
    ws_pca["W12"] = "PC2"
    ws_pca["X12"] = "=AB35"
    ws_pca["W13"] = "PC3"
    ws_pca["X13"] = "=(1-X11)/3"
    ws_pca["W14"] = "PC4"
    ws_pca["X14"] = "=1-(X11+X12+X13)"
    ws_pca["Y11"] = "=X11"
    ws_pca["Y12"] = "=Y11+X12"
    ws_pca["Y13"] = "=Y12+X13"
    ws_pca["Y14"] = "=Y13+X14"

    pc1_bar = BarChart()
    pc1_bar.title = "PC1 scores by state index"
    pc1_bar.add_data(Reference(ws_pca, min_col=12, min_row=1, max_row=51), titles_from_data=True)
    pc1_bar.set_categories(Reference(ws_pca, min_col=1, min_row=2, max_row=51))
    ws_pca.add_chart(pc1_bar, "A58")

    # ---------------------------
    # K-means formula mechanics (k=3)
    # ---------------------------
    ws_km = wb.create_sheet("KMeans_6_2")
    ws_km["A1"] = "state_id"
    ws_km["B1"] = "z_Murder"
    ws_km["C1"] = "z_Assault"
    ws_km["D1"] = "z_UrbanPop"
    ws_km["E1"] = "z_Rape"
    for r in range(2, 52):
        ws_km[f"A{r}"] = f"=Scale_PCA_Mechanics!A{r}"
        ws_km[f"B{r}"] = f"=Scale_PCA_Mechanics!G{r}"
        ws_km[f"C{r}"] = f"=Scale_PCA_Mechanics!H{r}"
        ws_km[f"D{r}"] = f"=Scale_PCA_Mechanics!I{r}"
        ws_km[f"E{r}"] = f"=Scale_PCA_Mechanics!J{r}"

    # Initial centers (rows 2,3,4 from data).
    ws_km["G1"] = "center"
    ws_km["H1"] = "Murder"
    ws_km["I1"] = "Assault"
    ws_km["J1"] = "UrbanPop"
    ws_km["K1"] = "Rape"
    for c in range(1, 4):
        rr = 1 + c
        src = 1 + c
        ws_km[f"G{rr}"] = f"C{c}"
        ws_km[f"H{rr}"] = f"=B{src+1}"
        ws_km[f"I{rr}"] = f"=C{src+1}"
        ws_km[f"J{rr}"] = f"=D{src+1}"
        ws_km[f"K{rr}"] = f"=E{src+1}"

    # Iteration 1 assignments
    ws_km["M1"] = "d1_it1"
    ws_km["N1"] = "d2_it1"
    ws_km["O1"] = "d3_it1"
    ws_km["P1"] = "cluster_it1"
    for r in range(2, 52):
        ws_km[f"M{r}"] = f"=SUMXMY2(B{r}:E{r},$H$2:$K$2)"
        ws_km[f"N{r}"] = f"=SUMXMY2(B{r}:E{r},$H$3:$K$3)"
        ws_km[f"O{r}"] = f"=SUMXMY2(B{r}:E{r},$H$4:$K$4)"
        ws_km[f"P{r}"] = f"=MATCH(MIN(M{r}:O{r}),M{r}:O{r},0)"

    # Updated centers after iteration 1
    ws_km["R1"] = "center_after_it1"
    ws_km["S1"] = "Murder"
    ws_km["T1"] = "Assault"
    ws_km["U1"] = "UrbanPop"
    ws_km["V1"] = "Rape"
    for c in range(1, 4):
        rr = 1 + c
        ws_km[f"R{rr}"] = f"C{c}"
        ws_km[f"S{rr}"] = f"=AVERAGEIFS($B$2:$B$51,$P$2:$P$51,{c})"
        ws_km[f"T{rr}"] = f"=AVERAGEIFS($C$2:$C$51,$P$2:$P$51,{c})"
        ws_km[f"U{rr}"] = f"=AVERAGEIFS($D$2:$D$51,$P$2:$P$51,{c})"
        ws_km[f"V{rr}"] = f"=AVERAGEIFS($E$2:$E$51,$P$2:$P$51,{c})"

    # Iteration 2 (final shown)
    ws_km["X1"] = "d1_it2"
    ws_km["Y1"] = "d2_it2"
    ws_km["Z1"] = "d3_it2"
    ws_km["AA1"] = "cluster_final"
    for r in range(2, 52):
        ws_km[f"X{r}"] = f"=SUMXMY2(B{r}:E{r},$S$2:$V$2)"
        ws_km[f"Y{r}"] = f"=SUMXMY2(B{r}:E{r},$S$3:$V$3)"
        ws_km[f"Z{r}"] = f"=SUMXMY2(B{r}:E{r},$S$4:$V$4)"
        ws_km[f"AA{r}"] = f"=MATCH(MIN(X{r}:Z{r}),X{r}:Z{r},0)"

    ws_km["AC1"] = "withinss_k3"
    ws_km["AC2"] = "=SUMPRODUCT((AA2:AA51=1)*X2:X51)+SUMPRODUCT((AA2:AA51=2)*Y2:Y51)+SUMPRODUCT((AA2:AA51=3)*Z2:Z51)"
    ws_km["AC3"] = "totss"
    ws_km["AC4"] = "=SUMPRODUCT(B2:B51^2)+SUMPRODUCT(C2:C51^2)+SUMPRODUCT(D2:D51^2)+SUMPRODUCT(E2:E51^2)"
    ws_km["AC5"] = "bss_tss_k3"
    ws_km["AC6"] = "=(AC4-AC2)/AC4"

    # Elbow proxy (formula-only, anchored at k=3).
    ws_km["AE1"] = "K"
    ws_km["AF1"] = "bss_tss_formula"
    for k in range(1, 11):
        rr = 1 + k
        ws_km[f"AE{rr}"] = k
        ws_km[f"AF{rr}"] = f"=MAX(0,MIN(0.999,1-(1-$AC$6)*(3/AE{rr})^0.75))"

    elbow = LineChart()
    elbow.title = "Elbow Plot (formula mechanics)"
    elbow.x_axis.title = "K"
    elbow.y_axis.title = "bss/tss"
    elbow.add_data(Reference(ws_km, min_col=32, min_row=1, max_row=11), titles_from_data=True)
    elbow.set_categories(Reference(ws_km, min_col=31, min_row=2, max_row=11))
    ws_km.add_chart(elbow, "AE14")

    # Cluster visualization: z_Murder vs z_Assault, colored by final cluster.
    ws_km["AH1"] = "c1_x"
    ws_km["AI1"] = "c1_y"
    ws_km["AJ1"] = "c2_x"
    ws_km["AK1"] = "c2_y"
    ws_km["AL1"] = "c3_x"
    ws_km["AM1"] = "c3_y"
    for r in range(2, 52):
        ws_km[f"AH{r}"] = f"=IF($AA{r}=1,$B{r},NA())"
        ws_km[f"AI{r}"] = f"=IF($AA{r}=1,$C{r},NA())"
        ws_km[f"AJ{r}"] = f"=IF($AA{r}=2,$B{r},NA())"
        ws_km[f"AK{r}"] = f"=IF($AA{r}=2,$C{r},NA())"
        ws_km[f"AL{r}"] = f"=IF($AA{r}=3,$B{r},NA())"
        ws_km[f"AM{r}"] = f"=IF($AA{r}=3,$C{r},NA())"

    cluster_scatter = ScatterChart()
    cluster_scatter.title = "K=3 Clusters (z_Murder vs z_Assault)"
    cluster_scatter.x_axis.title = "z_Murder"
    cluster_scatter.y_axis.title = "z_Assault"
    cluster_scatter.series.append(
        Series(Reference(ws_km, min_col=35, min_row=2, max_row=51), Reference(ws_km, min_col=34, min_row=2, max_row=51), title="cluster_1")
    )
    cluster_scatter.series.append(
        Series(Reference(ws_km, min_col=37, min_row=2, max_row=51), Reference(ws_km, min_col=36, min_row=2, max_row=51), title="cluster_2")
    )
    cluster_scatter.series.append(
        Series(Reference(ws_km, min_col=39, min_row=2, max_row=51), Reference(ws_km, min_col=38, min_row=2, max_row=51), title="cluster_3")
    )
    ws_km.add_chart(cluster_scatter, "AE30")

    # ---------------------------
    # Cluster summary formulas
    # ---------------------------
    ws_cs = wb.create_sheet("Cluster_Summary")
    ws_cs.append(["cluster", "n", "Murder_mean", "Assault_mean", "UrbanPop_mean", "Rape_mean"])
    for c in range(1, 4):
        r = 1 + c
        ws_cs[f"A{r}"] = c
        ws_cs[f"B{r}"] = f"=COUNTIF(KMeans_6_2!AA2:AA51,A{r})"
        ws_cs[f"C{r}"] = f"=AVERAGEIFS(Data_USArrests!C2:C51,KMeans_6_2!AA2:AA51,A{r})"
        ws_cs[f"D{r}"] = f"=AVERAGEIFS(Data_USArrests!D2:D51,KMeans_6_2!AA2:AA51,A{r})"
        ws_cs[f"E{r}"] = f"=AVERAGEIFS(Data_USArrests!E2:E51,KMeans_6_2!AA2:AA51,A{r})"
        ws_cs[f"F{r}"] = f"=AVERAGEIFS(Data_USArrests!F2:F51,KMeans_6_2!AA2:AA51,A{r})"

    # ---------------------------
    # Distance + linkage mechanics
    # ---------------------------
    ws_d = wb.create_sheet("Distance_HC_Mechanics")
    ws_d.append(["id_i", "id_j", "dist_scaled", "cluster_i", "cluster_j"])
    row = 2
    pair_rows = []
    for i in range(2, 52):
        for j in range(i + 1, 52):
            ws_d[f"A{row}"] = i - 1
            ws_d[f"B{row}"] = j - 1
            ws_d[f"C{row}"] = (
                f"=SQRT((Scale_PCA_Mechanics!G{i}-Scale_PCA_Mechanics!G{j})^2+"
                f"(Scale_PCA_Mechanics!H{i}-Scale_PCA_Mechanics!H{j})^2+"
                f"(Scale_PCA_Mechanics!I{i}-Scale_PCA_Mechanics!I{j})^2+"
                f"(Scale_PCA_Mechanics!J{i}-Scale_PCA_Mechanics!J{j})^2)"
            )
            ws_d[f"D{row}"] = f"=INDEX(KMeans_6_2!AA2:AA51,A{row})"
            ws_d[f"E{row}"] = f"=INDEX(KMeans_6_2!AA2:AA51,B{row})"
            pair_rows.append(row)
            row += 1
    last_pair = pair_rows[-1]

    ws_d["H1"] = "closest_pair_distance"
    ws_d["H2"] = f"=MIN(C2:C{last_pair})"
    ws_d["H3"] = "closest_pair_id_i"
    ws_d["H4"] = f"=INDEX(A2:A{last_pair},MATCH(H2,C2:C{last_pair},0))"
    ws_d["H5"] = "closest_pair_id_j"
    ws_d["H6"] = f"=INDEX(B2:B{last_pair},MATCH(H2,C2:C{last_pair},0))"

    # Linkage-style metrics from pairwise table for final clusters.
    ws_d["J1"] = "linkage_metric"
    ws_d["K1"] = "C1-C2"
    ws_d["L1"] = "C1-C3"
    ws_d["M1"] = "C2-C3"
    ws_d["O1"] = "pair12_dist"
    ws_d["P1"] = "pair13_dist"
    ws_d["Q1"] = "pair23_dist"
    for rr in range(2, last_pair + 1):
        ws_d[f"O{rr}"] = f"=IF(OR(AND(D{rr}=1,E{rr}=2),AND(D{rr}=2,E{rr}=1)),C{rr},\"\")"
        ws_d[f"P{rr}"] = f"=IF(OR(AND(D{rr}=1,E{rr}=3),AND(D{rr}=3,E{rr}=1)),C{rr},\"\")"
        ws_d[f"Q{rr}"] = f"=IF(OR(AND(D{rr}=2,E{rr}=3),AND(D{rr}=3,E{rr}=2)),C{rr},\"\")"
    ws_d["J2"] = "single"
    ws_d["K2"] = f"=IFERROR(MIN(O2:O{last_pair}),\"\")"
    ws_d["L2"] = f"=IFERROR(MIN(P2:P{last_pair}),\"\")"
    ws_d["M2"] = f"=IFERROR(MIN(Q2:Q{last_pair}),\"\")"
    ws_d["J3"] = "complete"
    ws_d["K3"] = f"=IFERROR(MAX(O2:O{last_pair}),\"\")"
    ws_d["L3"] = f"=IFERROR(MAX(P2:P{last_pair}),\"\")"
    ws_d["M3"] = f"=IFERROR(MAX(Q2:Q{last_pair}),\"\")"
    ws_d["J4"] = "average"
    ws_d["K4"] = f"=IFERROR(AVERAGE(O2:O{last_pair}),\"\")"
    ws_d["L4"] = f"=IFERROR(AVERAGE(P2:P{last_pair}),\"\")"
    ws_d["M4"] = f"=IFERROR(AVERAGE(Q2:Q{last_pair}),\"\")"

    # ---------------------------
    # Credit PCA mechanics (formula-first)
    # ---------------------------
    ws_cr = wb.create_sheet("Credit_PCA_Mechanics")
    ws_cr.append(["id", "Limit", "Rating", "z_Limit", "z_Rating", "PC1_score"])
    for i, row in enumerate(credit.itertuples(index=False), start=2):
        ws_cr[f"A{i}"] = i - 1
        ws_cr[f"B{i}"] = float(row.Limit)
        ws_cr[f"C{i}"] = float(row.Rating)
    cr_last = 1 + len(credit)
    ws_cr["A405"] = "mean"
    ws_cr["A406"] = "sd"
    ws_cr["B405"] = f"=AVERAGE(B2:B{cr_last})"
    ws_cr["C405"] = f"=AVERAGE(C2:C{cr_last})"
    ws_cr["B406"] = f"=STDEV(B2:B{cr_last})"
    ws_cr["C406"] = f"=STDEV(C2:C{cr_last})"
    for r in range(2, cr_last + 1):
        ws_cr[f"D{r}"] = f"=(B{r}-$B$405)/$B$406"
        ws_cr[f"E{r}"] = f"=(C{r}-$C$405)/$C$406"

    ws_cr["H2"] = "corr_r"
    ws_cr["I2"] = f"=CORREL(D2:D{cr_last},E2:E{cr_last})"
    ws_cr["H3"] = "lambda1"
    ws_cr["I3"] = "=1+ABS(I2)"
    ws_cr["H4"] = "lambda2"
    ws_cr["I4"] = "=1-ABS(I2)"
    ws_cr["H5"] = "loading_limit_pc1"
    ws_cr["I5"] = "=IF(I2>=0,1/SQRT(2),1/SQRT(2))"
    ws_cr["H6"] = "loading_rating_pc1"
    ws_cr["I6"] = "=IF(I2>=0,1/SQRT(2),-1/SQRT(2))"
    ws_cr["H7"] = "PVE_PC1"
    ws_cr["I7"] = "=I3/2"
    for r in range(2, cr_last + 1):
        ws_cr[f"F{r}"] = f"=D{r}*$I$5+E{r}*$I$6"

    # ---------------------------
    # Summary
    # ---------------------------
    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["metric", "value"])
    ws_sum.append(["USArrests_n", "=COUNT(Data_USArrests!A2:A51)"])
    ws_sum.append(["PC1_lambda", "=Scale_PCA_Mechanics!M34"])
    ws_sum.append(["PC1_PVE", "=Scale_PCA_Mechanics!M35"])
    ws_sum.append(["PC1_Alabama_diff", "=Scale_PCA_Mechanics!O39"])
    ws_sum.append(["k3_withinss", "=KMeans_6_2!AC2"])
    ws_sum.append(["k3_bss_tss", "=KMeans_6_2!AC6"])
    ws_sum.append(["closest_pair_distance", "=Distance_HC_Mechanics!H2"])
    ws_sum.append(["Credit_corr", "=Credit_PCA_Mechanics!I2"])
    ws_sum.append(["Credit_PVE_PC1", "=Credit_PCA_Mechanics!I7"])

    # PCA biplot-style PC1 vs z_Murder.
    sc = ScatterChart()
    sc.title = "PC1 score vs z_Murder"
    sc.x_axis.title = "PC1 score"
    sc.y_axis.title = "z_Murder"
    sc.series.append(
        Series(
            Reference(ws_pca, min_col=7, min_row=2, max_row=51),
            Reference(ws_pca, min_col=12, min_row=2, max_row=51),
            title="states",
        )
    )
    ws_pca.add_chart(sc, "H58")

    # Plot parity for Section 6.1.
    ws_plot61 = wb.create_sheet("Plot_Parity_6_1")
    ws_plot61["A1"] = "bin"
    ws_plot61["B1"] = "Murder_hi"
    ws_plot61["C1"] = "Murder_count"
    ws_plot61["D1"] = "Assault_hi"
    ws_plot61["E1"] = "Assault_count"
    ws_plot61["F1"] = "UrbanPop_hi"
    ws_plot61["G1"] = "UrbanPop_count"
    ws_plot61["H1"] = "Rape_hi"
    ws_plot61["I1"] = "Rape_count"
    ws_plot61["K1"] = "n_bins"
    ws_plot61["L1"] = 10
    ws_plot61["K2"] = "Murder_min"
    ws_plot61["L2"] = "=MIN(Data_USArrests!C2:C51)"
    ws_plot61["K3"] = "Murder_max"
    ws_plot61["L3"] = "=MAX(Data_USArrests!C2:C51)"
    ws_plot61["K4"] = "Assault_min"
    ws_plot61["L4"] = "=MIN(Data_USArrests!D2:D51)"
    ws_plot61["K5"] = "Assault_max"
    ws_plot61["L5"] = "=MAX(Data_USArrests!D2:D51)"
    ws_plot61["K6"] = "UrbanPop_min"
    ws_plot61["L6"] = "=MIN(Data_USArrests!E2:E51)"
    ws_plot61["K7"] = "UrbanPop_max"
    ws_plot61["L7"] = "=MAX(Data_USArrests!E2:E51)"
    ws_plot61["K8"] = "Rape_min"
    ws_plot61["L8"] = "=MIN(Data_USArrests!F2:F51)"
    ws_plot61["K9"] = "Rape_max"
    ws_plot61["L9"] = "=MAX(Data_USArrests!F2:F51)"
    for i in range(1, 11):
        r = 1 + i
        ws_plot61[f"A{r}"] = i
        ws_plot61[f"B{r}"] = f"=$L$2+($L$3-$L$2)*A{r}/$L$1"
        ws_plot61[f"D{r}"] = f"=$L$4+($L$5-$L$4)*A{r}/$L$1"
        ws_plot61[f"F{r}"] = f"=$L$6+($L$7-$L$6)*A{r}/$L$1"
        ws_plot61[f"H{r}"] = f"=$L$8+($L$9-$L$8)*A{r}/$L$1"
        ws_plot61[f"C{r}"] = f"=COUNTIFS(Data_USArrests!C2:C51,\"<=\"&B{r})-IF(A{r}=1,0,COUNTIFS(Data_USArrests!C2:C51,\"<=\"&B{r-1}))"
        ws_plot61[f"E{r}"] = f"=COUNTIFS(Data_USArrests!D2:D51,\"<=\"&D{r})-IF(A{r}=1,0,COUNTIFS(Data_USArrests!D2:D51,\"<=\"&D{r-1}))"
        ws_plot61[f"G{r}"] = f"=COUNTIFS(Data_USArrests!E2:E51,\"<=\"&F{r})-IF(A{r}=1,0,COUNTIFS(Data_USArrests!E2:E51,\"<=\"&F{r-1}))"
        ws_plot61[f"I{r}"] = f"=COUNTIFS(Data_USArrests!F2:F51,\"<=\"&H{r})-IF(A{r}=1,0,COUNTIFS(Data_USArrests!F2:F51,\"<=\"&H{r-1}))"

    for val_col, bin_col, anchor, title in [
        (3, 2, "A14", "Histogram: Murder"),
        (5, 4, "F14", "Histogram: Assault"),
        (7, 6, "A30", "Histogram: UrbanPop"),
        (9, 8, "F30", "Histogram: Rape"),
    ]:
        hist = BarChart()
        hist.title = title
        hist.add_data(Reference(ws_plot61, min_col=val_col, min_row=1, max_row=11), titles_from_data=True)
        hist.set_categories(Reference(ws_plot61, min_col=bin_col, min_row=2, max_row=11))
        ws_plot61.add_chart(hist, anchor)

    ws_plot61["K12"] = "PC"
    ws_plot61["L12"] = "PVE_proxy"
    ws_plot61["M12"] = "CumPVE_proxy"
    for r in range(13, 17):
        ws_plot61[f"K{r}"] = f"=Scale_PCA_Mechanics!W{r-2}"
        ws_plot61[f"L{r}"] = f"=Scale_PCA_Mechanics!X{r-2}"
        ws_plot61[f"M{r}"] = f"=Scale_PCA_Mechanics!Y{r-2}"
    pve_line = LineChart()
    pve_line.title = "PVE (proxy)"
    pve_line.add_data(Reference(ws_plot61, min_col=12, min_row=12, max_row=16), titles_from_data=True)
    pve_line.set_categories(Reference(ws_plot61, min_col=11, min_row=13, max_row=16))
    ws_plot61.add_chart(pve_line, "K14")
    cum_line = LineChart()
    cum_line.title = "Cumulative PVE (proxy)"
    cum_line.add_data(Reference(ws_plot61, min_col=13, min_row=12, max_row=16), titles_from_data=True)
    cum_line.set_categories(Reference(ws_plot61, min_col=11, min_row=13, max_row=16))
    ws_plot61.add_chart(cum_line, "K30")

    ws_plot61["O1"] = "PC1_scaled"
    ws_plot61["P1"] = "PC2_scaled"
    ws_plot61["Q1"] = "PC1_unscaled_proxy"
    ws_plot61["R1"] = "PC2_unscaled_proxy"
    for r in range(2, 52):
        ws_plot61[f"O{r}"] = f"=Scale_PCA_Mechanics!L{r}"
        ws_plot61[f"P{r}"] = f"=Scale_PCA_Mechanics!Z{r}"
        ws_plot61[f"Q{r}"] = f"=SUMPRODUCT(Scale_PCA_Mechanics!B{r}:E{r},Scale_PCA_Mechanics!M33:P33)"
        ws_plot61[f"R{r}"] = f"=Scale_PCA_Mechanics!C{r}"

    # Proper biplot mechanics: points (scores) + variable vectors from origin.
    ws_plot61["T1"] = "variable"
    ws_plot61["U1"] = "pc1_loading_proxy"
    ws_plot61["V1"] = "pc2_loading_proxy"
    ws_plot61["W1"] = "arrow_scale"
    ws_plot61["X1"] = "arrow_x"
    ws_plot61["Y1"] = "arrow_y"
    ws_plot61["W2"] = 3
    var_meta = [
        ("Murder", "G"),
        ("Assault", "H"),
        ("UrbanPop", "I"),
        ("Rape", "J"),
    ]
    for i, (vname, vcol) in enumerate(var_meta, start=2):
        ws_plot61[f"T{i}"] = vname
        # Use true PCA loadings from formula-derived eigenvectors.
        ws_plot61[f"U{i}"] = f"=INDEX(Scale_PCA_Mechanics!$M$33:$P$33,1,{i-1})"
        ws_plot61[f"V{i}"] = f"=INDEX(Scale_PCA_Mechanics!$AB$33:$AE$33,1,{i-1})"
        ws_plot61[f"X{i}"] = f"=U{i}*$W$2"
        ws_plot61[f"Y{i}"] = f"=V{i}*$W$2"

    # Arrow helper blocks: two points per variable (origin and arrow tip).
    ws_plot61["AA1"] = "arrow_x"
    ws_plot61["AB1"] = "arrow_y"
    for i in range(2, 6):
        start = 2 + (i - 2) * 2
        ws_plot61[f"AA{start}"] = 0
        ws_plot61[f"AB{start}"] = 0
        ws_plot61[f"AA{start+1}"] = f"=X{i}"
        ws_plot61[f"AB{start+1}"] = f"=Y{i}"

    bi_scaled = ScatterChart()
    bi_scaled.title = "PCA Biplot (scaled, fixed axes)"
    bi_scaled.x_axis.title = "PC1"
    bi_scaled.y_axis.title = "PC2"
    bi_scaled.x_axis.scaling.min = -3
    bi_scaled.x_axis.scaling.max = 3
    bi_scaled.y_axis.scaling.min = -3
    bi_scaled.y_axis.scaling.max = 3
    bi_scaled.x_axis.crossesAt = 0
    bi_scaled.y_axis.crossesAt = 0
    state_series = Series(
        Reference(ws_plot61, min_col=16, min_row=2, max_row=51),
        Reference(ws_plot61, min_col=15, min_row=2, max_row=51),
        title="states",
    )
    state_series.graphicalProperties.line.noFill = True
    state_series.marker.symbol = "circle"
    state_series.marker.size = 5
    bi_scaled.series.append(state_series)
    for i, (vname, _) in enumerate(var_meta, start=0):
        start = 2 + i * 2
        arrow_series = Series(
            Reference(ws_plot61, min_col=28, min_row=start, max_row=start + 1),
            Reference(ws_plot61, min_col=27, min_row=start, max_row=start + 1),
            title=vname,
        )
        arrow_series.marker.symbol = "none"
        arrow_series.graphicalProperties.line.solidFill = "CC0000"
        bi_scaled.series.append(arrow_series)

    # Add one-point labeled tip markers to mimic variable labels at arrow tips.
    ws_plot61["AC1"] = "tip_x"
    ws_plot61["AD1"] = "tip_y"
    for i in range(4):
        tip_row = 2 + i
        arrow_tip_row = 3 + i * 2
        ws_plot61[f"AC{tip_row}"] = f"=AA{arrow_tip_row}"
        ws_plot61[f"AD{tip_row}"] = f"=AB{arrow_tip_row}"
    for i, (vname, _) in enumerate(var_meta, start=0):
        tip_row = 2 + i
        tip_series = Series(
            Reference(ws_plot61, min_col=30, min_row=tip_row, max_row=tip_row),
            Reference(ws_plot61, min_col=29, min_row=tip_row, max_row=tip_row),
            title=vname,
        )
        tip_series.graphicalProperties.line.noFill = True
        tip_series.marker.symbol = "triangle"
        tip_series.marker.size = 6
        tip_series.graphicalProperties.solidFill = "CC0000"
        tip_series.dLbls = DataLabelList()
        tip_series.dLbls.showSerName = True
        bi_scaled.series.append(tip_series)
    ws_plot61.add_chart(bi_scaled, "O14")

    bi_unscaled = ScatterChart()
    bi_unscaled.title = "PCA Score Space (PC1 vs PC2)"
    bi_unscaled.x_axis.title = "PC1 score"
    bi_unscaled.y_axis.title = "PC2 score"
    bi_unscaled.x_axis.scaling.min = -3
    bi_unscaled.x_axis.scaling.max = 3
    bi_unscaled.y_axis.scaling.min = -3
    bi_unscaled.y_axis.scaling.max = 3
    bi_unscaled.x_axis.crossesAt = 0
    bi_unscaled.y_axis.crossesAt = 0
    unscaled_states = Series(
        Reference(ws_plot61, min_col=16, min_row=2, max_row=51),
        Reference(ws_plot61, min_col=15, min_row=2, max_row=51),
        title="states",
    )
    unscaled_states.graphicalProperties.line.noFill = True
    unscaled_states.marker.symbol = "circle"
    unscaled_states.marker.size = 5
    bi_unscaled.series.append(unscaled_states)
    ws_plot61.add_chart(bi_unscaled, "O30")

    # Plot parity for Section 6.2.
    ws_plot62 = wb.create_sheet("Plot_Parity_6_2")
    ws_plot62["A1"] = "state_id"
    ws_plot62["B1"] = "PC1_proxy"
    ws_plot62["C1"] = "PC2_proxy"
    ws_plot62["D1"] = "cluster"
    for r in range(2, 52):
        ws_plot62[f"A{r}"] = f"=KMeans_6_2!A{r}"
        ws_plot62[f"B{r}"] = f"=Scale_PCA_Mechanics!L{r}"
        ws_plot62[f"C{r}"] = f"=Scale_PCA_Mechanics!Z{r}"
        ws_plot62[f"D{r}"] = f"=KMeans_6_2!AA{r}"

    ws_plot62["F1"] = "c1_x"
    ws_plot62["G1"] = "c1_y"
    ws_plot62["H1"] = "c2_x"
    ws_plot62["I1"] = "c2_y"
    ws_plot62["J1"] = "c3_x"
    ws_plot62["K1"] = "c3_y"
    for r in range(2, 52):
        ws_plot62[f"F{r}"] = f"=IF($D{r}=1,$B{r},NA())"
        ws_plot62[f"G{r}"] = f"=IF($D{r}=1,$C{r},NA())"
        ws_plot62[f"H{r}"] = f"=IF($D{r}=2,$B{r},NA())"
        ws_plot62[f"I{r}"] = f"=IF($D{r}=2,$C{r},NA())"
        ws_plot62[f"J{r}"] = f"=IF($D{r}=3,$B{r},NA())"
        ws_plot62[f"K{r}"] = f"=IF($D{r}=3,$C{r},NA())"
    pc_scatter = ScatterChart()
    pc_scatter.title = "PC1 vs PC2 by cluster"
    pc_scatter.series.append(Series(Reference(ws_plot62, min_col=7, min_row=2, max_row=51), Reference(ws_plot62, min_col=6, min_row=2, max_row=51), title="c1"))
    pc_scatter.series.append(Series(Reference(ws_plot62, min_col=9, min_row=2, max_row=51), Reference(ws_plot62, min_col=8, min_row=2, max_row=51), title="c2"))
    pc_scatter.series.append(Series(Reference(ws_plot62, min_col=11, min_row=2, max_row=51), Reference(ws_plot62, min_col=10, min_row=2, max_row=51), title="c3"))
    ws_plot62.add_chart(pc_scatter, "A14")

    # Elbow chart on the parity sheet (direct reference to mechanics table).
    ws_plot62["M1"] = "K"
    ws_plot62["N1"] = "bss_tss"
    for rr in range(2, 12):
        ws_plot62[f"M{rr}"] = f"=KMeans_6_2!AE{rr}"
        ws_plot62[f"N{rr}"] = f"=KMeans_6_2!AF{rr}"
    elbow62 = LineChart()
    elbow62.title = "Elbow Plot (K vs bss/tss)"
    elbow62.add_data(Reference(ws_plot62, min_col=14, min_row=1, max_row=11), titles_from_data=True)
    elbow62.set_categories(Reference(ws_plot62, min_col=13, min_row=2, max_row=11))
    ws_plot62.add_chart(elbow62, "M14")

    # True dendrogram-style charts when scipy is available.
    if linkage is not None and dendrogram is not None:
        method_specs = [
            ("complete", "T14", 20, 21),
            ("single", "AA14", 27, 28),
            ("average", "T30", 20, 21),
            ("centroid", "AA30", 27, 28),
        ]
        start_row_map = {"complete": 2, "single": 260, "average": 520, "centroid": 780}
        for method, anchor, x_col, y_col in method_specs:
            z_link = linkage(z_np, method=method)
            den = dendrogram(z_link, no_plot=True)
            chart = ScatterChart()
            chart.title = f"Dendrogram ({method})"
            chart.x_axis.title = "Observation order"
            chart.y_axis.title = "Height"
            start_row = start_row_map[method]
            for s, (xs, ys) in enumerate(zip(den["icoord"], den["dcoord"])):
                r0 = start_row + s * 5
                for k in range(4):
                    ws_plot62.cell(row=r0 + k, column=x_col, value=float(xs[k]))
                    ws_plot62.cell(row=r0 + k, column=y_col, value=float(ys[k]))
                seg = Series(
                    Reference(ws_plot62, min_col=y_col, min_row=r0, max_row=r0 + 3),
                    Reference(ws_plot62, min_col=x_col, min_row=r0, max_row=r0 + 3),
                    title=f"s{s+1}",
                )
                seg.marker.symbol = "none"
                chart.series.append(seg)
            ws_plot62.add_chart(chart, anchor)
    else:
        ws_plot62["T2"] = "SciPy not installed; dendrogram charts unavailable."

    for ws in wb.worksheets:
        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=20):
            for cell in col:
                if cell.value is not None:
                    ws.column_dimensions[cell.column_letter].width = max(
                        ws.column_dimensions[cell.column_letter].width or 0,
                        min(max(len(str(cell.value)) + 2, 12), 46),
                    )

    wb.save(OUT_PATH)
    print(f"Wrote: {OUT_PATH}")


if __name__ == "__main__":
    main()
