from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.worksheet.formula import ArrayFormula


ROOT = Path(__file__).resolve().parent
OUT_PATH = ROOT / "asa_pa_4_3_interaction_glm_mechanics.xlsx"


def main() -> None:
    df = pd.read_csv(ROOT / "interactiondata.csv").copy()
    df.insert(0, "row_id", range(1, len(df) + 1))
    n = len(df)
    last = n + 1

    wb = Workbook()
    ws_readme = wb.active
    ws_readme.title = "README"
    ws_readme.append(["sheet", "purpose"])
    ws_readme.append(["Data", "Chunk 1 data, helper columns, and model predictions in formulas."])
    ws_readme.append(["Main_Effects_Mechanics", "Chunk 2: log-link main-effects mechanics via matrix normal equations on log(actual)."])
    ws_readme.append(["Interaction_Mechanics", "Chunk 3: interaction mechanics via matrix normal equations on log(actual)."])
    ws_readme.append(["Model_Comparison", "AIC-style comparison and SSE on log scale."])
    ws_readme.append(["Plot_Parity", "Formula-driven line charts matching chunk visuals."])
    ws_readme["A10"] = "All mechanics use Excel formulas only (no LINEST)."

    # ------------------------
    # Data + helper columns + predictions
    # ------------------------
    ws_d = wb.create_sheet("Data")
    headers = [
        "row_id",
        "age",
        "sex",
        "actual",
        "log_actual",
        "sex_M",
        "sex_age",
        "pred_main",
        "log_pred_main",
        "pred_interact",
        "log_pred_interact",
        "age_F",
        "log_actual_F",
        "log_pred_main_F",
        "log_pred_interact_F",
        "age_M",
        "log_actual_M",
        "log_pred_main_M",
        "log_pred_interact_M",
    ]
    ws_d.append(headers)
    for row in df[["row_id", "age", "sex", "actual"]].itertuples(index=False):
        ws_d.append(list(row) + [None] * (len(headers) - 4))

    for r in range(2, last + 1):
        ws_d[f"E{r}"] = f"=LN(D{r})"
        ws_d[f"F{r}"] = f'=--(C{r}="M")'
        ws_d[f"G{r}"] = f"=B{r}*F{r}"
        ws_d[f"H{r}"] = f"=EXP(Main_Effects_Mechanics!$G$11+Main_Effects_Mechanics!$G$12*F{r}+Main_Effects_Mechanics!$G$13*B{r})"
        ws_d[f"I{r}"] = f"=LN(H{r})"
        ws_d[f"J{r}"] = (
            f"=EXP(Interaction_Mechanics!$H$12+Interaction_Mechanics!$H$13*F{r}"
            f"+Interaction_Mechanics!$H$14*B{r}+Interaction_Mechanics!$H$15*G{r})"
        )
        ws_d[f"K{r}"] = f"=LN(J{r})"

        ws_d[f"L{r}"] = f'=IF(C{r}="F",B{r},NA())'
        ws_d[f"M{r}"] = f'=IF(C{r}="F",E{r},NA())'
        ws_d[f"N{r}"] = f'=IF(C{r}="F",I{r},NA())'
        ws_d[f"O{r}"] = f'=IF(C{r}="F",K{r},NA())'
        ws_d[f"P{r}"] = f'=IF(C{r}="M",B{r},NA())'
        ws_d[f"Q{r}"] = f'=IF(C{r}="M",E{r},NA())'
        ws_d[f"R{r}"] = f'=IF(C{r}="M",I{r},NA())'
        ws_d[f"S{r}"] = f'=IF(C{r}="M",K{r},NA())'

    # ------------------------
    # Main effects mechanics
    # ------------------------
    ws_m = wb.create_sheet("Main_Effects_Mechanics")
    ws_m["A1"] = "Chunk 2 model: log(actual) ~ 1 + sex_M + age"
    ws_m["A3"] = "X'X"
    ws_m["F3"] = "X'y (y=log_actual)"
    ws_m["B4"] = "Intercept"
    ws_m["C4"] = "sex_M"
    ws_m["D4"] = "age"
    ws_m["F4"] = "vector"

    ws_m["B5"] = f"=COUNT(Data!A2:A{last})"
    ws_m["C5"] = f"=SUM(Data!F2:F{last})"
    ws_m["D5"] = f"=SUM(Data!B2:B{last})"
    ws_m["B6"] = f"=SUM(Data!F2:F{last})"
    ws_m["C6"] = f"=SUMPRODUCT(Data!F2:F{last},Data!F2:F{last})"
    ws_m["D6"] = f"=SUMPRODUCT(Data!F2:F{last},Data!B2:B{last})"
    ws_m["B7"] = f"=SUM(Data!B2:B{last})"
    ws_m["C7"] = f"=SUMPRODUCT(Data!F2:F{last},Data!B2:B{last})"
    ws_m["D7"] = f"=SUMPRODUCT(Data!B2:B{last},Data!B2:B{last})"

    ws_m["F5"] = f"=SUM(Data!E2:E{last})"
    ws_m["F6"] = f"=SUMPRODUCT(Data!F2:F{last},Data!E2:E{last})"
    ws_m["F7"] = f"=SUMPRODUCT(Data!B2:B{last},Data!E2:E{last})"

    ws_m["B10"] = "inv(X'X)"
    ws_m["B11"] = ArrayFormula("B11:D13", "=MINVERSE(B5:D7)")

    ws_m["G10"] = "beta_main"
    ws_m["G11"] = "=MMULT(B11:D13,F5:F7)"
    ws_m["G12"] = "=MMULT(B12:D12,F5:F7)"
    ws_m["G13"] = "=MMULT(B13:D13,F5:F7)"

    ws_m["A16"] = "SSE_log_main"
    ws_m["B16"] = f"=SUMXMY2(Data!E2:E{last},Data!I2:I{last})"

    # ------------------------
    # Interaction mechanics
    # ------------------------
    ws_i = wb.create_sheet("Interaction_Mechanics")
    ws_i["A1"] = "Chunk 3 model: log(actual) ~ 1 + sex_M + age + sex_M*age"
    ws_i["A3"] = "X'X"
    ws_i["G3"] = "X'y (y=log_actual)"
    ws_i["B4"] = "Intercept"
    ws_i["C4"] = "sex_M"
    ws_i["D4"] = "age"
    ws_i["E4"] = "sex_age"
    ws_i["G4"] = "vector"

    # X'X entries
    ws_i["B5"] = f"=COUNT(Data!A2:A{last})"
    ws_i["C5"] = f"=SUM(Data!F2:F{last})"
    ws_i["D5"] = f"=SUM(Data!B2:B{last})"
    ws_i["E5"] = f"=SUM(Data!G2:G{last})"
    ws_i["B6"] = f"=SUM(Data!F2:F{last})"
    ws_i["C6"] = f"=SUMPRODUCT(Data!F2:F{last},Data!F2:F{last})"
    ws_i["D6"] = f"=SUMPRODUCT(Data!F2:F{last},Data!B2:B{last})"
    ws_i["E6"] = f"=SUMPRODUCT(Data!F2:F{last},Data!G2:G{last})"
    ws_i["B7"] = f"=SUM(Data!B2:B{last})"
    ws_i["C7"] = f"=SUMPRODUCT(Data!B2:B{last},Data!F2:F{last})"
    ws_i["D7"] = f"=SUMPRODUCT(Data!B2:B{last},Data!B2:B{last})"
    ws_i["E7"] = f"=SUMPRODUCT(Data!B2:B{last},Data!G2:G{last})"
    ws_i["B8"] = f"=SUM(Data!G2:G{last})"
    ws_i["C8"] = f"=SUMPRODUCT(Data!G2:G{last},Data!F2:F{last})"
    ws_i["D8"] = f"=SUMPRODUCT(Data!G2:G{last},Data!B2:B{last})"
    ws_i["E8"] = f"=SUMPRODUCT(Data!G2:G{last},Data!G2:G{last})"

    # X'y
    ws_i["G5"] = f"=SUM(Data!E2:E{last})"
    ws_i["G6"] = f"=SUMPRODUCT(Data!F2:F{last},Data!E2:E{last})"
    ws_i["G7"] = f"=SUMPRODUCT(Data!B2:B{last},Data!E2:E{last})"
    ws_i["G8"] = f"=SUMPRODUCT(Data!G2:G{last},Data!E2:E{last})"

    ws_i["B11"] = "inv(X'X)"
    ws_i["B12"] = ArrayFormula("B12:E15", "=MINVERSE(B5:E8)")
    ws_i["H11"] = "beta_interact"
    ws_i["H12"] = "=MMULT(B12:E12,G5:G8)"
    ws_i["H13"] = "=MMULT(B13:E13,G5:G8)"
    ws_i["H14"] = "=MMULT(B14:E14,G5:G8)"
    ws_i["H15"] = "=MMULT(B15:E15,G5:G8)"

    ws_i["A18"] = "SSE_log_interact"
    ws_i["B18"] = f"=SUMXMY2(Data!E2:E{last},Data!K2:K{last})"

    # ------------------------
    # Comparison
    # ------------------------
    ws_c = wb.create_sheet("Model_Comparison")
    ws_c.append(["metric", "value"])
    ws_c.append(["n", f"=COUNT(Data!A2:A{last})"])
    ws_c.append(["k_main", 3])
    ws_c.append(["k_interact", 4])
    ws_c.append(["SSE_log_main", "=Main_Effects_Mechanics!B16"])
    ws_c.append(["SSE_log_interact", "=Interaction_Mechanics!B18"])
    ws_c.append(["AIC_proxy_main", "=B2*LN(B5/B2)+2*B3"])
    ws_c.append(["AIC_proxy_interact", "=B2*LN(B6/B2)+2*B4"])
    ws_c.append(["Delta_AIC_proxy (interact-main)", "=B8-B7"])

    # ------------------------
    # Charts parity
    # ------------------------
    ws_p = wb.create_sheet("Plot_Parity")
    ws_p["A1"] = "Chunk 1 parity: log(actual) vs age by sex"
    ws_p["A20"] = "Chunk 2 parity: log(pred_main_only) + log(actual)"
    ws_p["A39"] = "Chunk 3 parity: log(pred_interact) + log(actual)"

    def add_lines(anchor: str, title: str, y_cols: list[tuple[int, str]]) -> None:
        ch = LineChart()
        ch.title = title
        ch.x_axis.title = "age"
        ch.y_axis.title = "log(value)"
        for y_col, nm in y_cols:
            ser = Series(
                Reference(ws_d, min_col=y_col, min_row=2, max_row=last),
                Reference(ws_d, min_col=(12 if y_col in [13, 14, 15] else 16), min_row=2, max_row=last),
                title=nm,
            )
            ch.series.append(ser)
        ws_p.add_chart(ch, anchor)

    # Build charts with separate F and M series (XY scatter + lines)
    c1 = ScatterChart()
    c1.title = "log(actual) by sex"
    c1.x_axis.title = "age"
    c1.y_axis.title = "log(actual)"
    s = Series(Reference(ws_d, min_col=13, min_row=2, max_row=last), Reference(ws_d, min_col=12, min_row=2, max_row=last), title="F actual")
    s.marker.symbol = "none"
    c1.series.append(s)
    s = Series(Reference(ws_d, min_col=17, min_row=2, max_row=last), Reference(ws_d, min_col=16, min_row=2, max_row=last), title="M actual")
    s.marker.symbol = "none"
    c1.series.append(s)
    ws_p.add_chart(c1, "A3")

    c2 = ScatterChart()
    c2.title = "Main-only: log(pred) vs log(actual)"
    c2.x_axis.title = "age"
    c2.y_axis.title = "log(value)"
    for y_col, x_col, name in [
        (14, 12, "F pred_main"),
        (13, 12, "F actual"),
        (18, 16, "M pred_main"),
        (17, 16, "M actual"),
    ]:
        s = Series(Reference(ws_d, min_col=y_col, min_row=2, max_row=last), Reference(ws_d, min_col=x_col, min_row=2, max_row=last), title=name)
        s.marker.symbol = "none"
        c2.series.append(s)
    ws_p.add_chart(c2, "A22")

    c3 = ScatterChart()
    c3.title = "Interaction: log(pred) vs log(actual)"
    c3.x_axis.title = "age"
    c3.y_axis.title = "log(value)"
    for y_col, x_col, name in [
        (15, 12, "F pred_interact"),
        (13, 12, "F actual"),
        (19, 16, "M pred_interact"),
        (17, 16, "M actual"),
    ]:
        s = Series(Reference(ws_d, min_col=y_col, min_row=2, max_row=last), Reference(ws_d, min_col=x_col, min_row=2, max_row=last), title=name)
        s.marker.symbol = "none"
        c3.series.append(s)
    ws_p.add_chart(c3, "A41")

    # Widths
    for ws in wb.worksheets:
        for col in ws.iter_cols(min_col=1, max_col=30, min_row=1, max_row=1):
            for cell in col:
                if cell.value is not None:
                    ws.column_dimensions[cell.column_letter].width = max(
                        ws.column_dimensions[cell.column_letter].width or 0,
                        min(max(len(str(cell.value)) + 2, 12), 50),
                    )

    wb.save(OUT_PATH)
    print(f"Wrote: {OUT_PATH}")


if __name__ == "__main__":
    main()
