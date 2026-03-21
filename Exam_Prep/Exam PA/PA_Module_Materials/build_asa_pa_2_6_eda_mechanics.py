from __future__ import annotations

import math
from pathlib import Path

import pandas as pd
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series


ROOT = Path(__file__).resolve().parent
OUT_PATH = ROOT / "asa_pa_2_6_eda_mechanics.xlsx"


def main() -> None:
    m = pd.read_csv(ROOT / "soa_mortality_data.csv")
    keep = [
        "prodcat",
        "issstate",
        "distchan",
        "smoker",
        "sex",
        "issage",
        "uwkey",
        "year",
        "resind_ind",
        "actual_cnt",
        "actual_face",
        "duration",
        "uwtype",
    ]
    m = m[keep].copy()
    m.insert(0, "row_id", range(1, len(m) + 1))
    m["actual_face_pos"] = m["actual_face"].where(m["actual_face"] > 0)
    m["log_actual_face_pos"] = m["actual_face_pos"].map(
        lambda x: None if pd.isna(x) else float(math.log10(x))
    )

    wb = Workbook()
    ws_readme = wb.active
    ws_readme.title = "README"
    ws_readme.append(["sheet", "purpose"])
    ws_readme.append(["Mortality_Data", "Chunk 1 source data and helper columns (positive-face / log-face)."])
    ws_readme.append(["Boxplot_Mechanics", "Chunks 2-4 five-number summaries used by split boxplots."])
    ws_readme.append(["Histogram_Mechanics", "Chunks 5-7 stacked/split histogram bin mechanics."])
    ws_readme.append(["Bar_Mechanics", "Chunk 8 stacked/split bar count mechanics."])
    ws_readme.append(["Scatter_Mechanics", "Chunks 9-12 scatter mechanics and subset sampling tables."])
    ws_readme["A10"] = "Formula-first EDA parity workbook for asa_pa_2.6.rmd chunks 1-12."

    # ------------------------
    # Data
    # ------------------------
    ws_d = wb.create_sheet("Mortality_Data")
    cols = [
        "row_id",
        *keep,
        "actual_face_pos",
        "log_actual_face_pos",
    ]
    ws_d.append(cols)
    for row in m[cols].itertuples(index=False):
        ws_d.append(list(row))
    last = 1 + len(m)

    # ------------------------
    # Boxplot mechanics
    # ------------------------
    ws_b = wb.create_sheet("Boxplot_Mechanics")
    ws_b["A1"] = "Chunk 2: issstate vs issage"
    ws_b["A2"] = "issstate"
    ws_b["B2"] = "n"
    ws_b["C2"] = "min"
    ws_b["D2"] = "q1"
    ws_b["E2"] = "median"
    ws_b["F2"] = "q3"
    ws_b["G2"] = "max"
    states = sorted(m["issstate"].dropna().astype(str).unique().tolist())[:25]
    r0 = 3
    for i, st in enumerate(states, start=r0):
        ws_b[f"A{i}"] = st
        ws_b[f"B{i}"] = f'=COUNTIFS(Mortality_Data!C2:C{last},A{i})'
        ws_b[f"C{i}"] = ArrayFormula(
            f"C{i}",
            f"=MIN(IF(Mortality_Data!C2:C{last}=A{i},Mortality_Data!G2:G{last}))",
        )
        ws_b[f"D{i}"] = ArrayFormula(
            f"D{i}",
            f"=QUARTILE(IF(Mortality_Data!C2:C{last}=A{i},Mortality_Data!G2:G{last}),1)",
        )
        ws_b[f"E{i}"] = ArrayFormula(
            f"E{i}",
            f"=MEDIAN(IF(Mortality_Data!C2:C{last}=A{i},Mortality_Data!G2:G{last}))",
        )
        ws_b[f"F{i}"] = ArrayFormula(
            f"F{i}",
            f"=QUARTILE(IF(Mortality_Data!C2:C{last}=A{i},Mortality_Data!G2:G{last}),3)",
        )
        ws_b[f"G{i}"] = ArrayFormula(
            f"G{i}",
            f"=MAX(IF(Mortality_Data!C2:C{last}=A{i},Mortality_Data!G2:G{last}))",
        )

    ws_b["I1"] = "Chunk 3: year/sex/smoker summaries"
    ws_b["I2"] = "group"
    ws_b["J2"] = "n"
    ws_b["K2"] = "mean_issage"
    groups = [("year", "I"), ("resind_ind", "J"), ("sex", "F"), ("smoker", "E")]
    rr = 3
    for gname, col in groups:
        vals = sorted(m[gname].dropna().astype(str).unique().tolist())[:10]
        for v in vals:
            ws_b[f"I{rr}"] = f"{gname}:{v}"
            ws_b[f"J{rr}"] = f'=COUNTIFS(Mortality_Data!{col}2:{col}{last},"{v}")'
            ws_b[f"K{rr}"] = f'=AVERAGEIFS(Mortality_Data!G2:G{last},Mortality_Data!{col}2:{col}{last},"{v}")'
            rr += 1

    ws_b["M1"] = "Chunk 4: actual_face (actual_cnt>=1)"
    ws_b["M2"] = "group"
    ws_b["N2"] = "n_pos"
    ws_b["O2"] = "median_log_face"
    grp2 = [("resind_ind", "J"), ("uwkey", "H"), ("sex", "F"), ("smoker", "E")]
    rr = 3
    for gname, col in grp2:
        vals = sorted(m[gname].dropna().astype(str).unique().tolist())[:12]
        for v in vals:
            ws_b[f"M{rr}"] = f"{gname}:{v}"
            ws_b[f"N{rr}"] = f'=COUNTIFS(Mortality_Data!{col}2:{col}{last},"{v}",Mortality_Data!K2:K{last},">0")'
            ws_b[f"O{rr}"] = ArrayFormula(
                f"O{rr}",
                f'=MEDIAN(IF((Mortality_Data!{col}2:{col}{last}="{v}")*(Mortality_Data!K2:K{last}>0),Mortality_Data!P2:P{last}))',
            )
            rr += 1

    # ------------------------
    # Histogram mechanics
    # ------------------------
    ws_h = wb.create_sheet("Histogram_Mechanics")
    ws_h["A1"] = "bin_start"
    ws_h["B1"] = "bin_end"
    bins = list(range(0, 101, 5))
    for i, b in enumerate(bins[:-1], start=2):
        ws_h[f"A{i}"] = b
        ws_h[f"B{i}"] = bins[i - 1]

    ws_h["D1"] = "Chunk 5: prodcat histogram counts"
    prodcats = sorted(m["prodcat"].dropna().astype(str).unique().tolist())[:6]
    for j, pc in enumerate(prodcats, start=4):
        col = chr(64 + j)
        ws_h[f"{col}1"] = pc
        for i in range(2, 2 + len(bins) - 1):
            ws_h[f"{col}{i}"] = f'=COUNTIFS(Mortality_Data!G2:G{last},">="&$A{i},Mortality_Data!G2:G{last},"<"&$B{i},Mortality_Data!B2:B{last},"{pc}")'

    ws_h["L1"] = "Chunk 7: smoker histogram counts"
    for j, sm in enumerate(["N", "S", "U"], start=12):
        col = chr(64 + j)
        ws_h[f"{col}1"] = sm
        for i in range(2, 2 + len(bins) - 1):
            ws_h[f"{col}{i}"] = f'=COUNTIFS(Mortality_Data!G2:G{last},">="&$A{i},Mortality_Data!G2:G{last},"<"&$B{i},Mortality_Data!E2:E{last},"{sm}")'

    h_chart = LineChart()
    h_chart.title = "Issue age histogram by prodcat (counts)"
    h_chart.add_data(Reference(ws_h, min_col=4, min_row=1, max_col=3 + len(prodcats), max_row=1 + len(bins) - 1), titles_from_data=True)
    h_chart.set_categories(Reference(ws_h, min_col=1, min_row=2, max_row=1 + len(bins) - 1))
    ws_h.add_chart(h_chart, "R2")

    # ------------------------
    # Bar mechanics
    # ------------------------
    ws_bar = wb.create_sheet("Bar_Mechanics")
    ws_bar["A1"] = "Chunk 8: prodcat x smoker"
    ws_bar["A2"] = "prodcat"
    ws_bar["B2"] = "N"
    ws_bar["C2"] = "S"
    ws_bar["D2"] = "U"
    for i, pc in enumerate(prodcats, start=3):
        ws_bar[f"A{i}"] = pc
        ws_bar[f"B{i}"] = f'=COUNTIFS(Mortality_Data!B2:B{last},A{i},Mortality_Data!D2:D{last},"N")'
        ws_bar[f"C{i}"] = f'=COUNTIFS(Mortality_Data!B2:B{last},A{i},Mortality_Data!D2:D{last},"S")'
        ws_bar[f"D{i}"] = f'=COUNTIFS(Mortality_Data!B2:B{last},A{i},Mortality_Data!D2:D{last},"U")'

    ws_bar["F1"] = "Chunk 8: uwtype x uwkey (top uwkeys)"
    ws_bar["F2"] = "uwtype"
    top_uwkey = m["uwkey"].astype(str).value_counts().head(4).index.tolist()
    for j, uk in enumerate(top_uwkey, start=7):
        ws_bar.cell(2, j, uk)
    uwtypes = sorted(m["uwtype"].dropna().astype(str).unique().tolist())[:6]
    for i, ut in enumerate(uwtypes, start=3):
        ws_bar[f"F{i}"] = ut
        for j, uk in enumerate(top_uwkey, start=7):
            col = chr(64 + j)
            ws_bar[f"{col}{i}"] = f'=COUNTIFS(Mortality_Data!N2:N{last},$F{i},Mortality_Data!H2:H{last},"{uk}")'

    b1 = BarChart()
    b1.title = "prodcat by smoker counts"
    b1.add_data(Reference(ws_bar, min_col=2, min_row=2, max_col=4, max_row=2 + len(prodcats)), titles_from_data=True)
    b1.set_categories(Reference(ws_bar, min_col=1, min_row=3, max_row=2 + len(prodcats)))
    ws_bar.add_chart(b1, "A12")

    # ------------------------
    # Scatter mechanics (chunks 9-12)
    # ------------------------
    ws_s = wb.create_sheet("Scatter_Mechanics")
    ws_s["A1"] = "Sampled scatter points (every 50th row) for chart performance"
    ws_s.append(["row_id", "issage", "duration", "actual_face", "log_face_pos", "smoker", "sex", "prodcat", "resind_ind"])
    rows = m.iloc[::50, :].copy()
    for row in rows.itertuples(index=False):
        ws_s.append([row.row_id, row.issage, row.duration, row.actual_face, row.log_actual_face_pos, row.smoker, row.sex, row.prodcat, row.resind_ind])
    s_last = 1 + len(rows)

    # subsets for CHUNK 12 TRM
    ws_s["K1"] = "TRM subset sample"
    ws_s["K2"] = "duration"
    ws_s["L2"] = "log_face_pos"
    ws_s["M2"] = "smoker"
    ws_s["N2"] = "sex"
    rr = 3
    for r in range(2, s_last + 1):
        ws_s[f"K{rr}"] = f'=IF(H{r}="TRM",C{r},"")'
        ws_s[f"L{rr}"] = f'=IF(H{r}="TRM",E{r},"")'
        ws_s[f"M{rr}"] = f'=IF(H{r}="TRM",F{r},"")'
        ws_s[f"N{rr}"] = f'=IF(H{r}="TRM",G{r},"")'
        rr += 1

    sc1 = ScatterChart()
    sc1.title = "Chunk 9: issage vs duration"
    sc1.x_axis.title = "issage"
    sc1.y_axis.title = "duration"
    sc1.series.append(Series(Reference(ws_s, min_col=3, min_row=2, max_row=s_last), Reference(ws_s, min_col=2, min_row=2, max_row=s_last), title="all"))
    ws_s.add_chart(sc1, "A14")

    sc2 = ScatterChart()
    sc2.title = "Chunk 10/11: duration vs log(actual_face>0)"
    sc2.x_axis.title = "duration"
    sc2.y_axis.title = "log10(actual_face)"
    sc2.series.append(Series(Reference(ws_s, min_col=5, min_row=2, max_row=s_last), Reference(ws_s, min_col=3, min_row=2, max_row=s_last), title="all posface"))
    ws_s.add_chart(sc2, "I14")

    # widths
    for ws in wb.worksheets:
        for col in ws.iter_cols(min_col=1, max_col=30, min_row=1, max_row=1):
            for cell in col:
                if cell.value is not None:
                    ws.column_dimensions[cell.column_letter].width = max(
                        ws.column_dimensions[cell.column_letter].width or 0,
                        min(max(len(str(cell.value)) + 2, 12), 48),
                    )

    wb.save(OUT_PATH)
    print(f"Wrote: {OUT_PATH}")


if __name__ == "__main__":
    main()
