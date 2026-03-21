from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula


ROOT = Path(__file__).resolve().parent
OUT_PATH = ROOT / "asa_pa_4_2_glm_mechanics.xlsx"


def main() -> None:
    df = pd.read_csv(ROOT / "insurance.csv")
    df = df.copy()
    df.insert(0, "row_id", range(1, len(df) + 1))
    df["District_factor"] = df["District"].astype(str)
    # Keep source data raw; AvgClaims is computed in-sheet with formulas.

    district_levels = sorted(df["District_factor"].unique().tolist(), key=lambda x: int(x))
    group_levels = df["Group"].drop_duplicates().tolist()
    age_levels = df["Age"].drop_duplicates().tolist()
    n = len(df)
    data_last = n + 1

    wb = Workbook()
    ws_readme = wb.active
    ws_readme.title = "README"
    ws_readme.append(["sheet", "purpose"])
    ws_readme.append(["Insurance_Data", "Chunk 1 data load, factor recode, and AvgClaims helper."])
    ws_readme.append(["Design_Matrix", "Dummy-variable design matrix used by all model mechanics."])
    ws_readme.append(["Poisson_No_Offset", "Chunk 2 mod1: Poisson without offset, gradient mechanics in formulas."])
    ws_readme.append(["Poisson_With_Offset", "Chunk 2 mod2: Poisson with log(Holders) offset, formula-only mechanics."])
    ws_readme.append(["Gaussian_Models", "Chunk 3 mod3/mod4/mod5 OLS and weighted least squares via matrix formulas."])
    ws_readme.append(["Model_Comparison", "SSE comparison matching sse1..sse5 with charts."])
    ws_readme["A10"] = "All calculations are performed with Excel formulas."

    # ------------------------
    # Chunk 1 data
    # ------------------------
    ws_data = wb.create_sheet("Insurance_Data")
    cols = [
        "row_id",
        "District_raw",
        "District_factor",
        "Group",
        "Age",
        "Holders",
        "Claims",
        "AvgClaims",
    ]
    ws_data.append(cols)
    for row in df[["row_id", "District", "District_factor", "Group", "Age", "Holders", "Claims"]].itertuples(index=False):
        ws_data.append(list(row) + [None])
    for r in range(2, data_last + 1):
        ws_data[f"H{r}"] = f"=IFERROR(G{r}/F{r},0)"

    # ------------------------
    # Design matrix (intercept + dummies)
    # ------------------------
    ws_x = wb.create_sheet("Design_Matrix")
    ws_x["A1"] = "row_id"
    ws_x["B1"] = "Intercept"

    x_headers = ["District_" + lvl for lvl in district_levels[1:]] + [
        "Group_" + lvl for lvl in group_levels[1:]
    ] + ["Age_" + lvl for lvl in age_levels[1:]]
    for i, h in enumerate(x_headers, start=3):
        ws_x.cell(1, i, h)
    ws_x["L1"] = "Claims"
    ws_x["M1"] = "Holders"
    ws_x["N1"] = "AvgClaims"

    for r in range(2, data_last + 1):
        ws_x[f"A{r}"] = f"=Insurance_Data!A{r}"
        ws_x[f"B{r}"] = 1
        c = 3
        for lvl in district_levels[1:]:
            ws_x.cell(r, c, f'=--(Insurance_Data!C{r}="{lvl}")')
            c += 1
        for lvl in group_levels[1:]:
            ws_x.cell(r, c, f'=--(Insurance_Data!D{r}="{lvl}")')
            c += 1
        for lvl in age_levels[1:]:
            ws_x.cell(r, c, f'=--(Insurance_Data!E{r}="{lvl}")')
            c += 1
        ws_x[f"L{r}"] = f"=Insurance_Data!G{r}"
        ws_x[f"M{r}"] = f"=Insurance_Data!F{r}"
        ws_x[f"N{r}"] = f"=Insurance_Data!H{r}"

    # ------------------------
    # Poisson mechanics helper builder
    # ------------------------
    coef_names = ["Intercept"] + x_headers
    n_coef = len(coef_names)
    n_iter = 8

    def build_poisson_sheet(sheet_name: str, with_offset: bool) -> None:
        ws = wb.create_sheet(sheet_name)
        ws["A1"] = "learning_rate"
        ws["B1"] = 0.00001
        ws["A2"] = "iterations"
        ws["B2"] = n_iter
        ws["A4"] = "coef"
        for j in range(n_iter + 1):
            ws.cell(4, 2 + j, f"iter_{j}")
        for i, nm in enumerate(coef_names, start=5):
            ws[f"A{i}"] = nm
            ws["B" + str(i)] = 0

        # Coefficient updates: beta_{k+1} = beta_k + lr * gradient_k / n
        grad_start = 74
        ws[f"A{grad_start - 1}"] = "Gradient block (score equations)"
        ws[f"A{grad_start}"] = "coef"
        for j in range(n_iter):
            ws.cell(grad_start, 2 + j, f"g_iter_{j}")
        for i, nm in enumerate(coef_names, start=grad_start + 1):
            ws[f"A{i}"] = nm

        for j in range(1, n_iter + 1):
            coef_col = get_column_letter(2 + j)
            prev_col = get_column_letter(1 + j)
            grad_col = get_column_letter(1 + j)
            for i in range(n_coef):
                rr = 5 + i
                gg = grad_start + 1 + i
                ws[f"{coef_col}{rr}"] = (
                    f"=IFERROR({prev_col}{rr}+$B$1*{grad_col}{gg}/COUNT(Design_Matrix!A2:A{data_last}),{prev_col}{rr})"
                )

        # Row-wise eta/mu by iteration
        ws["L4"] = "Row mechanics by iteration"
        for j in range(n_iter + 1):
            eta_col = get_column_letter(12 + 2 * j)
            mu_col = get_column_letter(13 + 2 * j)
            ws[f"{eta_col}5"] = f"eta_iter_{j}"
            ws[f"{mu_col}5"] = f"mu_iter_{j}"
            coef_col = get_column_letter(2 + j)
            for r in range(6, 6 + n):
                src = r - 4
                if with_offset:
                    ws[f"{eta_col}{r}"] = (
                        f"=IFERROR(SUMPRODUCT(Design_Matrix!B{src}:K{src},${coef_col}$5:${coef_col}$14),0)"
                        f"+IFERROR(LN(Design_Matrix!M{src}),0)"
                    )
                else:
                    ws[f"{eta_col}{r}"] = (
                        f"=IFERROR(SUMPRODUCT(Design_Matrix!B{src}:K{src},${coef_col}$5:${coef_col}$14),0)"
                    )
                # Clamp eta to avoid EXP overflow causing #NUM! cascades.
                ws[f"{mu_col}{r}"] = f"=EXP(MAX(-50,MIN(50,{eta_col}{r})))"

        # Gradients per iteration and coefficient
        for j in range(n_iter):
            mu_col = get_column_letter(13 + 2 * j)
            grad_col = get_column_letter(2 + j)
            for i in range(n_coef):
                rr = grad_start + 1 + i
                x_col = get_column_letter(2 + i)  # Design_Matrix B..K
                ws[f"{grad_col}{rr}"] = (
                    f"=IFERROR(SUMPRODUCT(Design_Matrix!{x_col}2:{x_col}{data_last},Design_Matrix!L2:L{data_last})"
                    f"-SUMPRODUCT(Design_Matrix!{x_col}2:{x_col}{data_last},{mu_col}6:{mu_col}{5+n}),0)"
                )

        final_mu_col = get_column_letter(13 + 2 * n_iter)
        ws["A90"] = "SSE"
        ws["B90"] = f"=SUMXMY2(Design_Matrix!L2:L{data_last},{final_mu_col}6:{final_mu_col}{5+n})"
        ws["A91"] = "mean_pred"
        ws["B91"] = f"=AVERAGE({final_mu_col}6:{final_mu_col}{5+n})"

        # Tiny chart to show convergence of SSE proxy by iteration
        ws["D90"] = "iter"
        ws["E90"] = "sse_iter"
        for j in range(n_iter + 1):
            row = 91 + j
            mu_col = get_column_letter(13 + 2 * j)
            ws[f"D{row}"] = j
            ws[f"E{row}"] = f"=SUMXMY2(Design_Matrix!L2:L{data_last},{mu_col}6:{mu_col}{5+n})"
        ch = LineChart()
        ch.title = f"{sheet_name} SSE by iteration"
        ch.add_data(Reference(ws, min_col=5, min_row=90, max_row=91 + n_iter), titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=4, min_row=91, max_row=91 + n_iter))
        ws.add_chart(ch, "G90")

    build_poisson_sheet("Poisson_No_Offset", with_offset=False)
    build_poisson_sheet("Poisson_With_Offset", with_offset=True)

    # ------------------------
    # Gaussian models via matrix formulas (mod3/mod4/mod5)
    # ------------------------
    ws_g = wb.create_sheet("Gaussian_Models")
    ws_g["A1"] = "mod3/mod4/mod5 mechanics (matrix formulas)"
    ws_g["A2"] = "mod3 (gaussian on Claims)"
    ws_g["M2"] = "mod4 (gaussian on AvgClaims)"
    ws_g["AB2"] = "mod5 (weighted gaussian on AvgClaims, weights=Holders)"

    # X'X and inverse for unweighted models
    ws_g["A4"] = "XTX"
    ws_g["A16"] = "inv(XTX)"
    ws_g["L4"] = "XTy_claims"
    ws_g["M4"] = "beta_mod3"
    ws_g["N4"] = "XTy_avgclaims"
    ws_g["O4"] = "beta_mod4"
    ws_g["A5"] = ArrayFormula(
        "A5:J14",
        f"=MMULT(TRANSPOSE(Design_Matrix!B2:K{data_last}),Design_Matrix!B2:K{data_last})",
    )
    ws_g["A16"] = ArrayFormula("A16:J25", "=MINVERSE(A5:J14)")
    ws_g["L5"] = ArrayFormula(
        "L5:L14",
        f"=MMULT(TRANSPOSE(Design_Matrix!B2:K{data_last}),Design_Matrix!L2:L{data_last})",
    )
    ws_g["N5"] = ArrayFormula(
        "N5:N14",
        f"=MMULT(TRANSPOSE(Design_Matrix!B2:K{data_last}),Design_Matrix!N2:N{data_last})",
    )
    # Coefficients via normal equations (no LINEST): beta = inv(X'X) * X'y
    ws_g["M5"] = ArrayFormula("M5:M14", "=MMULT(MINVERSE(A5:J14),L5:L14)")
    ws_g["O5"] = ArrayFormula("O5:O14", "=MMULT(MINVERSE(A5:J14),N5:N14)")

    ws_g["A28"] = "row_id"
    ws_g["B28"] = "pred3"
    ws_g["C28"] = "pred4_avgclaims"
    ws_g["D28"] = "pred4_claims"
    for r in range(29, 29 + n):
        src = r - 27
        ws_g[f"A{r}"] = f"=Design_Matrix!A{src}"
        ws_g[f"B{r}"] = f"=MMULT(Design_Matrix!B{src}:K{src},$M$5:$M$14)"
        ws_g[f"C{r}"] = f"=MMULT(Design_Matrix!B{src}:K{src},$O$5:$O$14)"
        ws_g[f"D{r}"] = f"=Design_Matrix!M{src}*C{r}"

    # Weighted mod5
    ws_g["Q4"] = "Weighted X = X*sqrt(w)"
    ws_g["AA4"] = "Weighted y = AvgClaims*sqrt(w)"
    for r in range(5, 5 + n):
        src = r - 3
        for j in range(10):
            col = get_column_letter(17 + j)  # Q..Z
            x_col = get_column_letter(2 + j)  # B..K
            ws_g[f"{col}{r}"] = f"=Design_Matrix!{x_col}{src}*SQRT(Design_Matrix!M{src})"
        ws_g[f"AA{r}"] = f"=Design_Matrix!N{src}*SQRT(Design_Matrix!M{src})"

    # Keep weighted row-level data in Q:AA and place weighted normal-equation
    # matrices in a separate non-overlapping block.
    ws_g["AC4"] = "X'WX"
    ws_g["AC16"] = ArrayFormula("AC16:AL25", f"=MMULT(TRANSPOSE(Q5:Z{4+n}),Q5:Z{4+n})")
    ws_g["AC27"] = ArrayFormula("AC27:AL36", "=MINVERSE(AC16:AL25)")
    ws_g["AM4"] = "X'Wy"
    # X'Wy as scalar formulas (avoids array-eval compatibility issues).
    for j in range(10):
        rr = 16 + j
        c = get_column_letter(17 + j)  # Q..Z
        ws_g[f"AM{rr}"] = f"=SUMPRODUCT({c}5:{c}{4+n},$AA$5:$AA${4+n})"
    ws_g["AM26"] = "beta_mod5"
    ws_g["AM27"] = ArrayFormula("AM27:AM36", "=MMULT(MINVERSE(AC16:AL25),AM16:AM25)")

    ws_g["F28"] = "pred5_avgclaims"
    ws_g["G28"] = "pred5_claims"
    for r in range(29, 29 + n):
        src = r - 27
        ws_g[f"F{r}"] = f"=MMULT(Design_Matrix!B{src}:K{src},$AM$27:$AM$36)"
        ws_g[f"G{r}"] = f"=Design_Matrix!M{src}*F{r}"

    ws_g["A95"] = "sse3"
    ws_g["B95"] = f"=SUMXMY2(Design_Matrix!L2:L{data_last},B29:B{28+n})"
    ws_g["A96"] = "sse4"
    ws_g["B96"] = f"=SUMXMY2(Design_Matrix!L2:L{data_last},D29:D{28+n})"
    ws_g["A97"] = "sse5"
    ws_g["B97"] = f"=SUMXMY2(Design_Matrix!L2:L{data_last},G29:G{28+n})"

    # ------------------------
    # Comparison summary + chart (sse1..sse5)
    # ------------------------
    ws_cmp = wb.create_sheet("Model_Comparison")
    ws_cmp.append(["Model", "SSE", "Source"])
    ws_cmp.append(["sse1_mod1_poisson_no_offset", "=Poisson_No_Offset!B90", "Chunk 2"])
    ws_cmp.append(["sse2_mod2_poisson_with_offset", "=Poisson_With_Offset!B90", "Chunk 2"])
    ws_cmp.append(["sse3_mod3_gaussian_claims", "=Gaussian_Models!B95", "Chunk 3"])
    ws_cmp.append(["sse4_mod4_gaussian_avgclaims", "=Gaussian_Models!B96", "Chunk 3"])
    ws_cmp.append(["sse5_mod5_weighted_gaussian", "=Gaussian_Models!B97", "Chunk 3"])
    ws_cmp["A8"] = "Expected from the Rmd: offset helps Poisson, weighted gaussian helps among gaussian fits."

    sse_chart = BarChart()
    sse_chart.title = "SSE comparison (sse1..sse5)"
    sse_chart.add_data(Reference(ws_cmp, min_col=2, min_row=1, max_row=6), titles_from_data=True)
    sse_chart.set_categories(Reference(ws_cmp, min_col=1, min_row=2, max_row=6))
    ws_cmp.add_chart(sse_chart, "E2")

    # Widths
    for ws in wb.worksheets:
        for col in ws.iter_cols(min_col=1, max_col=40, min_row=1, max_row=1):
            for cell in col:
                if cell.value is not None:
                    ws.column_dimensions[cell.column_letter].width = max(
                        ws.column_dimensions[cell.column_letter].width or 0,
                        min(max(len(str(cell.value)) + 2, 12), 50),
                    )

    wb.save(OUT_PATH)
    print(f"Wrote: {OUT_PATH}")


if __name__ == "__main__":
    main()
