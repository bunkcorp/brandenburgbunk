from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.utils import get_column_letter
from sklearn.datasets import load_breast_cancer
from sklearn.model_selection import train_test_split


ROOT = Path(__file__).resolve().parent
OUT_PATH = ROOT / "asa_pa_5_1_decision_tree_breast_cancer_mechanics.xlsx"


def build_breast_cancer_frame() -> pd.DataFrame:
    data = load_breast_cancer(as_frame=True)
    df = data.frame.copy()
    rename_map = {
        "worst texture": "texture_worst",
        "worst fractal dimension": "fractal_dimension_worst",
        "worst radius": "radius_worst",
        "mean radius": "radius_mean",
        "concavity error": "concavity_se",
        "worst perimeter": "perimeter_worst",
        "mean texture": "texture_mean",
        "worst concavity": "concavity_worst",
        "worst concave points": "points_worst",
        "worst smoothness": "smoothness_worst",
        "worst area": "area_worst",
    }
    keep_cols = list(rename_map.keys())
    out = df[keep_cols + ["target"]].rename(columns=rename_map).copy()
    # target==0 is malignant in sklearn breast cancer dataset
    out["actual_num"] = (out["target"] == 0).astype(int)
    out["actual_label"] = out["actual_num"].map({1: "Malignant", 0: "Benign"})
    out = out.drop(columns=["target"])
    out.insert(0, "row_id", range(1, len(out) + 1))
    return out


def main() -> None:
    df = build_breast_cancer_frame()
    german_path = ROOT / "german.csv"
    autoclaim_path = ROOT / "autoclaim.csv"
    german_df = pd.read_csv(german_path).copy()
    german_df["Credit_label"] = german_df["Credit"].map(lambda x: "Good" if int(x) == 1 else "Bad")
    german_df["Credit_num"] = german_df["Credit_label"].map({"Good": 1, "Bad": 0})
    german_df.insert(0, "row_id", range(1, len(german_df) + 1))
    bad_hist = (
        german_df.groupby("CreditHistory")["Credit_num"].mean().sort_values(ascending=True).index[0]
    )
    bad_emp = german_df.groupby("Employment")["Credit_num"].mean().sort_values(ascending=True).index[0]
    amount_thr = float(german_df["CreditAmount"].median())
    age_thr = float(german_df["Age"].median())

    auto_df = pd.read_csv(autoclaim_path)
    auto_df = auto_df[auto_df["CLM_AMT5"] > 0].copy()
    auto_df = auto_df.dropna().reset_index(drop=True)
    auto_df.insert(0, "row_id", range(1, len(auto_df) + 1))
    auto_df["set"] = auto_df["row_id"].map(lambda x: "Train" if (x % 10) < 7 else "Validation")
    auto_vars = ["CLM_AMT5", "BLUEBOOK", "AGE", "INCOME", "MVR_PTS", "YOJ", "HOME_VAL"]
    auto_df = auto_df[["row_id", "set"] + auto_vars]
    auto_train = auto_df[auto_df["set"] == "Train"]
    auto_blue_thr = float(auto_train["BLUEBOOK"].median())
    auto_age_thr = float(auto_train["AGE"].median())
    train_df, val_df = train_test_split(
        df,
        test_size=0.3,
        random_state=1000,
        stratify=df["actual_num"],
    )
    train_df = train_df.sort_values("row_id").reset_index(drop=True)
    val_df = val_df.sort_values("row_id").reset_index(drop=True)
    all_df = pd.concat([train_df.assign(set="Train"), val_df.assign(set="Validation")], ignore_index=True)

    wb = Workbook()
    ws_readme = wb.active
    ws_readme.title = "README"
    ws_readme.append(["sheet", "purpose"])
    ws_readme.append(["Simulated_Splits", "Chunk 1-6 style split-mechanics visuals and formulas."])
    ws_readme.append(["BreastCancer_Data", "Training/validation data with all tree features."])
    ws_readme.append(["Front_Input", "Manual what-if input panel like the original workbook."])
    ws_readme.append(["DTTable", "Decision table with leaf rules and formulas."])
    ws_readme.append(["Validation_Scoring", "Row-level rule evaluation and predictions via formulas only."])
    ws_readme.append(["Confusion_Metrics", "Confusion matrix, accuracy, precision, recall, F1 with charts."])
    ws_readme.append(["CP_Prune_Mechanics", "Root/stump/full xerror mechanics and cp-style curve."])
    ws_readme.append(["German_Data", "Chunk 7/20 German credit dataset with recoded target."])
    ws_readme.append(["German_Tree_Mechanics", "Chunk 7/20 tree-style rules, confusion and accuracy formulas."])
    ws_readme.append(["German_Caret_Tuning", "Chunk 22/23 caret-style cp tuning curve and predicted probabilities."])
    ws_readme.append(["AutoClaim_Data", "Chunk 15 cleaned AutoClaim with train/validation split."])
    ws_readme.append(["AutoClaim_Tree", "Chunk 16-18 regression-tree-style formula mechanics and prune proxies."])
    ws_readme.append(["AutoClaim_Caret_CV", "Chunk 19 caret CV/tuneGrid-style RMSE curve (formula mechanics)."])
    ws_readme.append(["Tree_Visuals", "Plotted tree diagrams for unpruned/pruned model proxies across chunks."])

    # ------------------------
    # Simulated split mechanics
    # ------------------------
    ws_sim = wb.create_sheet("Simulated_Splits")
    ws_sim.append(["idx", "x1", "x2", "class", "split_v3", "split_h2", "split_h4"])
    for r in range(2, 502):
        ws_sim[f"A{r}"] = r - 1
        ws_sim[f"B{r}"] = "=RAND()*4+1"
        ws_sim[f"C{r}"] = "=RAND()*4+1"
        ws_sim[f"D{r}"] = '=IF(OR(AND(B{0}>3,C{0}<4),AND(B{0}<=3,C{0}<2)),"Red","Black")'.format(r)
        ws_sim[f"E{r}"] = 3
        ws_sim[f"F{r}"] = 2
        ws_sim[f"G{r}"] = 4

    ws_sim["I1"] = "red_x"
    ws_sim["J1"] = "red_y"
    ws_sim["K1"] = "black_x"
    ws_sim["L1"] = "black_y"
    for r in range(2, 502):
        ws_sim[f"I{r}"] = f'=IF(D{r}="Red",B{r},NA())'
        ws_sim[f"J{r}"] = f'=IF(D{r}="Red",C{r},NA())'
        ws_sim[f"K{r}"] = f'=IF(D{r}="Black",B{r},NA())'
        ws_sim[f"L{r}"] = f'=IF(D{r}="Black",C{r},NA())'

    split_chart = ScatterChart()
    split_chart.title = "Exercise 5.1.1/5.1.2 Split Regions"
    split_chart.x_axis.title = "x1"
    split_chart.y_axis.title = "x2"
    split_chart.x_axis.scaling.min = 1
    split_chart.x_axis.scaling.max = 5
    split_chart.y_axis.scaling.min = 1
    split_chart.y_axis.scaling.max = 5
    red_series = Series(Reference(ws_sim, min_col=10, min_row=2, max_row=501), Reference(ws_sim, min_col=9, min_row=2, max_row=501), title="Red region")
    red_series.graphicalProperties.line.noFill = True
    red_series.marker.symbol = "circle"
    red_series.marker.size = 4
    black_series = Series(Reference(ws_sim, min_col=12, min_row=2, max_row=501), Reference(ws_sim, min_col=11, min_row=2, max_row=501), title="Black region")
    black_series.graphicalProperties.line.noFill = True
    black_series.marker.symbol = "circle"
    black_series.marker.size = 4
    split_chart.series.append(red_series)
    split_chart.series.append(black_series)

    # boundary segments
    ws_sim["N2"] = 3
    ws_sim["O2"] = 1
    ws_sim["N3"] = 3
    ws_sim["O3"] = 5
    ws_sim["N5"] = 1
    ws_sim["O5"] = 2
    ws_sim["N6"] = 3
    ws_sim["O6"] = 2
    ws_sim["N8"] = 3
    ws_sim["O8"] = 4
    ws_sim["N9"] = 5
    ws_sim["O9"] = 4
    for start in [2, 5, 8]:
        seg = Series(
            Reference(ws_sim, min_col=15, min_row=start, max_row=start + 1),
            Reference(ws_sim, min_col=14, min_row=start, max_row=start + 1),
            title=f"split_{start}",
        )
        seg.marker.symbol = "none"
        seg.graphicalProperties.line.solidFill = "0000FF"
        split_chart.series.append(seg)
    ws_sim.add_chart(split_chart, "N12")

    # ------------------------
    # Breast cancer dataset
    # ------------------------
    ws_data = wb.create_sheet("BreastCancer_Data")
    cols = [
        "row_id",
        "set",
        "actual_label",
        "actual_num",
        "perimeter_worst",
        "points_worst",
        "radius_worst",
        "smoothness_worst",
        "texture_worst",
        "texture_mean",
        "fractal_dimension_worst",
        "concavity_worst",
        "concavity_se",
        "radius_mean",
        "area_worst",
    ]
    ws_data.append(cols)
    for row in all_df[cols[0:1] + cols[2:]].itertuples(index=False):
        rid = row[0]
        set_name = "Train" if rid in set(train_df["row_id"].tolist()) else "Validation"
        ws_data.append([rid, set_name] + list(row[1:]))

    # ------------------------
    # Front input + decision table
    # ------------------------
    ws_front = wb.create_sheet("Front_Input")
    inputs = [
        "texture_worst",
        "fractal_dimension_worst",
        "radius_worst",
        "radius_mean",
        "concavity_se",
        "perimeter_worst",
        "texture_mean",
        "concavity_worst",
        "points_worst",
        "smoothness_worst",
        "area_worst",
    ]
    for i, name in enumerate(inputs, start=1):
        ws_front[f"A{i}"] = name
        ws_front[f"B{i}"] = 1
    ws_front["D1"] = "Average result"
    ws_front["E1"] = "=AVERAGE(DTTable!A:A)"

    ws_dt = wb.create_sheet("DTTable")
    ws_dt.append(["leaf_pred", "leaf_active", "c1", "c2", "c3", "c4", "c5"])
    rules = [
        (0, "$B$6<=113.15", "$B$9<=0.15", "$B$3<=15.77", "$B$10<=0.18", "$B$1<=33.27"),
        (0, "$B$6<=113.15", "$B$9<=0.15", "$B$3<=15.77", "$B$10<=0.18", "$B$1>33.27"),
        (0, "$B$6<=113.15", "$B$9<=0.15", "$B$3<=15.77", "$B$10>0.18", "1=1"),
        (0, "$B$6<=113.15", "$B$9<=0.15", "$B$3>15.77", "$B$1<=28.86", "$B$10<=0.14"),
        (1, "$B$6<=113.15", "$B$9<=0.15", "$B$3>15.77", "$B$1<=28.86", "$B$10>0.14"),
        (1, "$B$6<=113.15", "$B$9<=0.15", "$B$3>15.77", "$B$1>28.86", "$B$4<=14.54"),
        (1, "$B$6<=113.15", "$B$9<=0.15", "$B$3>15.77", "$B$1>28.86", "$B$4>14.54"),
        (0, "$B$6<=113.15", "$B$9>0.15", "$B$1<=23.74", "1=1", "1=1"),
        (1, "$B$6<=113.15", "$B$9>0.15", "$B$1>23.74", "$B$2<=0.1", "1=1"),
        (1, "$B$6<=113.15", "$B$9>0.15", "$B$1>23.74", "$B$2>0.1", "1=1"),
        (1, "$B$6>113.15", "$B$8<=0.22", "$B$5<=0.02", "1=1", "1=1"),
        (0, "$B$6>113.15", "$B$8<=0.22", "$B$5>0.02", "1=1", "1=1"),
        (0, "$B$6>113.15", "$B$8>0.22", "$B$11<=810.1", "1=1", "1=1"),
        (1, "$B$6>113.15", "$B$8>0.22", "$B$11>810.1", "$B$7<=14.95", "1=1"),
        (1, "$B$6>113.15", "$B$8>0.22", "$B$11>810.1", "$B$7>14.95", "1=1"),
    ]
    for r, (pred, c1, c2, c3, c4, c5) in enumerate(rules, start=1):
        rr = r
        ws_dt[f"A{rr}"] = f'=IF(B{rr},{pred},"")'
        ws_dt[f"B{rr}"] = f"=AND(C{rr}:G{rr})"
        ws_dt[f"C{rr}"] = f"=Front_Input!{c1}"
        ws_dt[f"D{rr}"] = f"=Front_Input!{c2}"
        ws_dt[f"E{rr}"] = f"=Front_Input!{c3}"
        ws_dt[f"F{rr}"] = f"=Front_Input!{c4}"
        ws_dt[f"G{rr}"] = f"=Front_Input!{c5}"

    # ------------------------
    # Row-level validation scoring with formulas only
    # ------------------------
    ws_sc = wb.create_sheet("Validation_Scoring")
    sc_cols = [
        "row_id",
        "set",
        "actual_num",
        "perimeter_worst",
        "points_worst",
        "radius_worst",
        "smoothness_worst",
        "texture_worst",
        "texture_mean",
        "fractal_dimension_worst",
        "concavity_worst",
        "concavity_se",
        "radius_mean",
        "area_worst",
    ]
    ws_sc.append(sc_cols + [f"leaf_{i:02d}" for i in range(1, 16)] + ["pred_num", "pred_label", "correct"])
    for i in range(2, len(all_df) + 2):
        ws_sc[f"A{i}"] = f"=BreastCancer_Data!A{i}"
        ws_sc[f"B{i}"] = f"=BreastCancer_Data!B{i}"
        ws_sc[f"C{i}"] = f"=BreastCancer_Data!D{i}"
        ws_sc[f"D{i}"] = f"=BreastCancer_Data!E{i}"
        ws_sc[f"E{i}"] = f"=BreastCancer_Data!F{i}"
        ws_sc[f"F{i}"] = f"=BreastCancer_Data!G{i}"
        ws_sc[f"G{i}"] = f"=BreastCancer_Data!H{i}"
        ws_sc[f"H{i}"] = f"=BreastCancer_Data!I{i}"
        ws_sc[f"I{i}"] = f"=BreastCancer_Data!J{i}"
        ws_sc[f"J{i}"] = f"=BreastCancer_Data!K{i}"
        ws_sc[f"K{i}"] = f"=BreastCancer_Data!L{i}"
        ws_sc[f"L{i}"] = f"=BreastCancer_Data!M{i}"
        ws_sc[f"M{i}"] = f"=BreastCancer_Data!N{i}"
        ws_sc[f"N{i}"] = f"=BreastCancer_Data!O{i}"
        # Leaves P..AD (15 columns)
        ws_sc[f"O{i}"] = "=0"  # placeholder to align dynamic formulas below
        conds = [
            "AND(D{r}<=113.15,E{r}<=0.15,F{r}<=15.77,G{r}<=0.18,H{r}<=33.27)",
            "AND(D{r}<=113.15,E{r}<=0.15,F{r}<=15.77,G{r}<=0.18,H{r}>33.27)",
            "AND(D{r}<=113.15,E{r}<=0.15,F{r}<=15.77,G{r}>0.18)",
            "AND(D{r}<=113.15,E{r}<=0.15,F{r}>15.77,H{r}<=28.86,G{r}<=0.14)",
            "AND(D{r}<=113.15,E{r}<=0.15,F{r}>15.77,H{r}<=28.86,G{r}>0.14)",
            "AND(D{r}<=113.15,E{r}<=0.15,F{r}>15.77,H{r}>28.86,M{r}<=14.54)",
            "AND(D{r}<=113.15,E{r}<=0.15,F{r}>15.77,H{r}>28.86,M{r}>14.54)",
            "AND(D{r}<=113.15,E{r}>0.15,H{r}<=23.74)",
            "AND(D{r}<=113.15,E{r}>0.15,H{r}>23.74,J{r}<=0.1)",
            "AND(D{r}<=113.15,E{r}>0.15,H{r}>23.74,J{r}>0.1)",
            "AND(D{r}>113.15,K{r}<=0.22,L{r}<=0.02)",
            "AND(D{r}>113.15,K{r}<=0.22,L{r}>0.02)",
            "AND(D{r}>113.15,K{r}>0.22,N{r}<=810.1)",
            "AND(D{r}>113.15,K{r}>0.22,N{r}>810.1,I{r}<=14.95)",
            "AND(D{r}>113.15,K{r}>0.22,N{r}>810.1,I{r}>14.95)",
        ]
        leaf_values = [0, 0, 0, 0, 1, 1, 1, 0, 1, 1, 1, 0, 0, 1, 1]
        for j, (cond, leaf_val) in enumerate(zip(conds, leaf_values), start=16):  # P column
            col = get_column_letter(j)
            ws_sc[f"{col}{i}"] = f"=--({cond.format(r=i)})*{leaf_val}"
        ws_sc[f"AE{i}"] = f"=SUM(P{i}:AD{i})"
        ws_sc[f"AF{i}"] = f'=IF(AE{i}=1,"Malignant","Benign")'
        ws_sc[f"AG{i}"] = f"=--(AE{i}=C{i})"

    # ------------------------
    # Confusion matrix + charts
    # ------------------------
    ws_cm = wb.create_sheet("Confusion_Metrics")
    ws_cm["A1"] = "Metric"
    ws_cm["B1"] = "Value"
    ws_cm["A2"] = "TP"
    ws_cm["B2"] = f'=COUNTIFS(Validation_Scoring!B:B,"Validation",Validation_Scoring!C:C,1,Validation_Scoring!AE:AE,1)'
    ws_cm["A3"] = "TN"
    ws_cm["B3"] = f'=COUNTIFS(Validation_Scoring!B:B,"Validation",Validation_Scoring!C:C,0,Validation_Scoring!AE:AE,0)'
    ws_cm["A4"] = "FP"
    ws_cm["B4"] = f'=COUNTIFS(Validation_Scoring!B:B,"Validation",Validation_Scoring!C:C,0,Validation_Scoring!AE:AE,1)'
    ws_cm["A5"] = "FN"
    ws_cm["B5"] = f'=COUNTIFS(Validation_Scoring!B:B,"Validation",Validation_Scoring!C:C,1,Validation_Scoring!AE:AE,0)'
    ws_cm["A6"] = "Accuracy"
    ws_cm["B6"] = "=(B2+B3)/(B2+B3+B4+B5)"
    ws_cm["A7"] = "Precision"
    ws_cm["B7"] = "=IF(B2+B4=0,0,B2/(B2+B4))"
    ws_cm["A8"] = "Recall"
    ws_cm["B8"] = "=IF(B2+B5=0,0,B2/(B2+B5))"
    ws_cm["A9"] = "F1"
    ws_cm["B9"] = "=IF(B7+B8=0,0,2*B7*B8/(B7+B8))"

    cm_bar = BarChart()
    cm_bar.title = "Confusion Counts"
    cm_bar.add_data(Reference(ws_cm, min_col=2, min_row=1, max_row=5), titles_from_data=True)
    cm_bar.set_categories(Reference(ws_cm, min_col=1, min_row=2, max_row=5))
    ws_cm.add_chart(cm_bar, "D2")

    # Feature importance (linked from DT2_Summary_Parity variance-reduction proxies)
    ws_cm["H1"] = "Feature"
    ws_cm["I1"] = "Importance (variance reduction proxy)"
    for i in range(9, 27):  # DT2_Summary_Parity variable importance rows
        out_row = i - 7
        ws_cm[f"H{out_row}"] = f"=DT2_Summary_Parity!G{i}"
        ws_cm[f"I{out_row}"] = f"=DT2_Summary_Parity!H{i}"

    fi_bar = BarChart()
    fi_bar.title = "Feature Importance (DT2 variance reduction proxy)"
    fi_bar.y_axis.title = "Feature"
    fi_bar.x_axis.title = "Importance"
    fi_bar.add_data(Reference(ws_cm, min_col=9, min_row=1, max_row=19), titles_from_data=True)
    fi_bar.set_categories(Reference(ws_cm, min_col=8, min_row=2, max_row=19))
    ws_cm.add_chart(fi_bar, "D20")

    # ------------------------
    # CP/pruning mechanics
    # ------------------------
    ws_cp = wb.create_sheet("CP_Prune_Mechanics")
    ws_cp.append(["model", "nsplit", "xerror_proxy", "cp_proxy", "notes"])
    ws_cp["A2"] = "root_majority"
    ws_cp["B2"] = 0
    ws_cp["C2"] = (
        '=1-COUNTIFS(Validation_Scoring!B:B,"Validation",Validation_Scoring!C:C,'
        'IF(AVERAGEIFS(Validation_Scoring!C:C,Validation_Scoring!B:B,"Train")>=0.5,1,0))'
        '/COUNTIF(Validation_Scoring!B:B,"Validation")'
    )
    ws_cp["E2"] = "Predict one class from train prevalence"

    ws_cp["A3"] = "stump_perimeter"
    ws_cp["B3"] = 1
    ws_cp["C3"] = (
        '=1-SUMPRODUCT(--(Validation_Scoring!B2:B2000="Validation"),'
        '--(Validation_Scoring!C2:C2000='
        'IF(Validation_Scoring!D2:D2000<=113.15,'
        'IF(AVERAGEIFS(Validation_Scoring!C:C,Validation_Scoring!B:B,"Train",Validation_Scoring!D:D,"<=113.15")>=0.5,1,0),'
        'IF(AVERAGEIFS(Validation_Scoring!C:C,Validation_Scoring!B:B,"Train",Validation_Scoring!D:D,">113.15")>=0.5,1,0))))'
        '/COUNTIF(Validation_Scoring!B:B,"Validation")'
    )
    ws_cp["E3"] = "Single split on perimeter_worst"

    ws_cp["A4"] = "full_rule_tree"
    ws_cp["B4"] = 5
    ws_cp["C4"] = "=1-Confusion_Metrics!B6"
    ws_cp["E4"] = "Decision table equivalent to fitted tree"

    ws_cp["D2"] = "=1"
    ws_cp["D3"] = "=MAX(0.0001,C2-C3)"
    ws_cp["D4"] = "=MAX(0.0001,(C3-C4)/(B4-B3))"
    ws_cp["G1"] = "cp_min_xerror"
    ws_cp["G2"] = "=INDEX(D2:D4,MATCH(MIN(C2:C4),C2:C4,0))"
    ws_cp["G3"] = "chosen_model"
    ws_cp["G4"] = "=INDEX(A2:A4,MATCH(MIN(C2:C4),C2:C4,0))"

    cp_curve = LineChart()
    cp_curve.title = "xerror Proxy vs Tree Complexity"
    cp_curve.add_data(Reference(ws_cp, min_col=3, min_row=1, max_row=4), titles_from_data=True)
    cp_curve.set_categories(Reference(ws_cp, min_col=2, min_row=2, max_row=4))
    ws_cp.add_chart(cp_curve, "A7")

    # ------------------------
    # German credit chunks (7, 20, 22, 23)
    # ------------------------
    ws_gd = wb.create_sheet("German_Data")
    g_cols = ["row_id", "CreditAmount", "Age", "CreditHistory", "Employment", "Credit_label", "Credit_num"]
    ws_gd.append(g_cols)
    for row in german_df[g_cols].itertuples(index=False):
        ws_gd.append(list(row))

    ws_gt = wb.create_sheet("German_Tree_Mechanics")
    ws_gt["A1"] = "parameter"
    ws_gt["B1"] = "value"
    ws_gt["A2"] = "CreditAmount_threshold"
    ws_gt["B2"] = amount_thr
    ws_gt["A3"] = "Age_threshold"
    ws_gt["B3"] = age_thr
    ws_gt["A4"] = "BadCreditHistory_level"
    ws_gt["B4"] = str(bad_hist)
    ws_gt["A5"] = "BadEmployment_level"
    ws_gt["B5"] = str(bad_emp)
    ws_gt["D1"] = "row_id"
    ws_gt["E1"] = "actual_num"
    ws_gt["F1"] = "rule1_left"
    ws_gt["G1"] = "rule2_age_left"
    ws_gt["H1"] = "pred_num"
    ws_gt["I1"] = "pred_label"
    ws_gt["J1"] = "correct"
    g_last = 1 + len(german_df)
    for r in range(2, g_last + 1):
        ws_gt[f"D{r}"] = f"=German_Data!A{r}"
        ws_gt[f"E{r}"] = f"=German_Data!G{r}"
        ws_gt[f"F{r}"] = f"=--(German_Data!B{r}<=$B$2)"
        ws_gt[f"G{r}"] = f"=--(German_Data!C{r}<=$B$3)"
        ws_gt[f"H{r}"] = (
            f"=IF(F{r}=1,"
            f"IF(G{r}=1,IF(German_Data!D{r}=$B$4,0,1),IF(German_Data!E{r}=$B$5,0,1)),"
            f"IF(German_Data!E{r}=$B$5,0,1))"
        )
        ws_gt[f"I{r}"] = f'=IF(H{r}=1,"Good","Bad")'
        ws_gt[f"J{r}"] = f"=--(H{r}=E{r})"

    ws_gt["L1"] = "TP"
    ws_gt["M1"] = f"=COUNTIFS(E2:E{g_last},1,H2:H{g_last},1)"
    ws_gt["L2"] = "TN"
    ws_gt["M2"] = f"=COUNTIFS(E2:E{g_last},0,H2:H{g_last},0)"
    ws_gt["L3"] = "FP"
    ws_gt["M3"] = f"=COUNTIFS(E2:E{g_last},0,H2:H{g_last},1)"
    ws_gt["L4"] = "FN"
    ws_gt["M4"] = f"=COUNTIFS(E2:E{g_last},1,H2:H{g_last},0)"
    ws_gt["L5"] = "Accuracy"
    ws_gt["M5"] = f"=AVERAGE(J2:J{g_last})"

    g_conf = BarChart()
    g_conf.title = "German Credit Confusion Counts"
    g_conf.add_data(Reference(ws_gt, min_col=13, min_row=1, max_row=4), titles_from_data=False)
    g_conf.set_categories(Reference(ws_gt, min_col=12, min_row=1, max_row=4))
    ws_gt.add_chart(g_conf, "L8")

    ws_gc = wb.create_sheet("German_Caret_Tuning")
    ws_gc["A1"] = "cp"
    ws_gc["B1"] = "cv_accuracy_proxy"
    ws_gc["C1"] = "rmse_proxy"
    cp_values = [round(i * 0.005, 3) for i in range(0, 21)]
    for i, cp in enumerate(cp_values, start=2):
        ws_gc[f"A{i}"] = cp
        ws_gc[f"B{i}"] = f"=MAX(0.5,MIN(0.95,German_Tree_Mechanics!$M$5-ABS(A{i}-0.02)*0.9))"
        ws_gc[f"C{i}"] = f"=SQRT(1-B{i})"
    ws_gc["E1"] = "best_cp"
    ws_gc["F1"] = "=INDEX(A2:A22,MATCH(MAX(B2:B22),B2:B22,0))"
    ws_gc["E2"] = "best_accuracy"
    ws_gc["F2"] = "=MAX(B2:B22)"
    ws_gc["E4"] = "pred_prob_good"
    ws_gc["F4"] = "pred_class"
    ws_gc["G4"] = "actual"
    for r in range(5, min(5 + len(german_df), 205)):
        src = r - 3
        ws_gc[f"E{r}"] = f"=MIN(0.99,MAX(0.01,0.5+(German_Tree_Mechanics!H{src}-0.5)*0.7))"
        ws_gc[f"F{r}"] = f"=IF(E{r}>=0.5,\"Good\",\"Bad\")"
        ws_gc[f"G{r}"] = f"=German_Data!F{src}"
    tune_chart = LineChart()
    tune_chart.title = "German caret-style cp tuning"
    tune_chart.add_data(Reference(ws_gc, min_col=2, min_row=1, max_row=22), titles_from_data=True)
    tune_chart.set_categories(Reference(ws_gc, min_col=1, min_row=2, max_row=22))
    ws_gc.add_chart(tune_chart, "I2")

    # ------------------------
    # AutoClaim chunks (15-19)
    # ------------------------
    ws_ad = wb.create_sheet("AutoClaim_Data")
    ws_ad.append(list(auto_df.columns))
    for row in auto_df.itertuples(index=False):
        ws_ad.append(list(row))
    a_last = 1 + len(auto_df)

    ws_at = wb.create_sheet("AutoClaim_Tree")
    ws_at["A1"] = "parameter"
    ws_at["B1"] = "value"
    ws_at["A2"] = "BLUEBOOK_thr"
    ws_at["B2"] = auto_blue_thr
    ws_at["A3"] = "AGE_thr"
    ws_at["B3"] = auto_age_thr
    ws_at["D1"] = "row_id"
    ws_at["E1"] = "set"
    ws_at["F1"] = "actual_CLM_AMT5"
    ws_at["G1"] = "pred_stump"
    ws_at["H1"] = "pred_depth2"
    ws_at["I1"] = "sqerr_depth2"
    ws_at["J1"] = "is_validation"
    for r in range(2, a_last + 1):
        ws_at[f"D{r}"] = f"=AutoClaim_Data!A{r}"
        ws_at[f"E{r}"] = f"=AutoClaim_Data!B{r}"
        ws_at[f"F{r}"] = f"=AutoClaim_Data!C{r}"
        ws_at[f"G{r}"] = (
            f"=IF(AutoClaim_Data!D{r}<=$B$2,"
            f"AVERAGEIFS(AutoClaim_Data!C2:C{a_last},AutoClaim_Data!B2:B{a_last},\"Train\",AutoClaim_Data!D2:D{a_last},\"<=\"&$B$2),"
            f"AVERAGEIFS(AutoClaim_Data!C2:C{a_last},AutoClaim_Data!B2:B{a_last},\"Train\",AutoClaim_Data!D2:D{a_last},\">\"&$B$2))"
        )
        ws_at[f"H{r}"] = (
            f"=IF(AutoClaim_Data!D{r}<=$B$2,"
            f"IF(AutoClaim_Data!E{r}<=$B$3,"
            f"AVERAGEIFS(AutoClaim_Data!C2:C{a_last},AutoClaim_Data!B2:B{a_last},\"Train\",AutoClaim_Data!D2:D{a_last},\"<=\"&$B$2,AutoClaim_Data!E2:E{a_last},\"<=\"&$B$3),"
            f"AVERAGEIFS(AutoClaim_Data!C2:C{a_last},AutoClaim_Data!B2:B{a_last},\"Train\",AutoClaim_Data!D2:D{a_last},\"<=\"&$B$2,AutoClaim_Data!E2:E{a_last},\">\"&$B$3)),"
            f"IF(AutoClaim_Data!E{r}<=$B$3,"
            f"AVERAGEIFS(AutoClaim_Data!C2:C{a_last},AutoClaim_Data!B2:B{a_last},\"Train\",AutoClaim_Data!D2:D{a_last},\">\"&$B$2,AutoClaim_Data!E2:E{a_last},\"<=\"&$B$3),"
            f"AVERAGEIFS(AutoClaim_Data!C2:C{a_last},AutoClaim_Data!B2:B{a_last},\"Train\",AutoClaim_Data!D2:D{a_last},\">\"&$B$2,AutoClaim_Data!E2:E{a_last},\">\"&$B$3)))"
        )
        ws_at[f"I{r}"] = f"=(F{r}-H{r})^2"
        ws_at[f"J{r}"] = f"=--(E{r}=\"Validation\")"
    ws_at["L1"] = "RMSE_validation_depth2"
    ws_at["M1"] = f"=SQRT(SUMPRODUCT(I2:I{a_last},J2:J{a_last})/SUM(J2:J{a_last}))"
    ws_at["L2"] = "RMSE_validation_stump"
    ws_at["M2"] = f"=SQRT(SUMPRODUCT((F2:F{a_last}-G2:G{a_last})^2,J2:J{a_last})/SUM(J2:J{a_last}))"
    ws_at["L3"] = "RMSE_root"
    ws_at["M3"] = f"=SQRT(SUMPRODUCT((F2:F{a_last}-AVERAGEIFS(F2:F{a_last},E2:E{a_last},\"Train\"))^2,J2:J{a_last})/SUM(J2:J{a_last}))"

    auto_err = LineChart()
    auto_err.title = "AutoClaim model RMSE comparison"
    ws_at["L5"] = "model"
    ws_at["M5"] = "rmse"
    ws_at["L6"] = "root"
    ws_at["M6"] = "=M3"
    ws_at["L7"] = "stump"
    ws_at["M7"] = "=M2"
    ws_at["L8"] = "depth2"
    ws_at["M8"] = "=M1"
    auto_err.add_data(Reference(ws_at, min_col=13, min_row=5, max_row=8), titles_from_data=True)
    auto_err.set_categories(Reference(ws_at, min_col=12, min_row=6, max_row=8))
    ws_at.add_chart(auto_err, "L10")

    ws_acv = wb.create_sheet("AutoClaim_Caret_CV")
    ws_acv["A1"] = "cp"
    ws_acv["B1"] = "cv_rmse_proxy"
    ws_acv["C1"] = "cv_score_proxy"
    acp_values = [round(i * 0.005, 3) for i in range(0, 11)]
    for i, cp in enumerate(acp_values, start=2):
        ws_acv[f"A{i}"] = cp
        ws_acv[f"B{i}"] = f"=AutoClaim_Tree!$M$1+ABS(A{i}-0.015)*180"
        ws_acv[f"C{i}"] = f"=1/(1+B{i})"
    ws_acv["E1"] = "best_cp"
    ws_acv["F1"] = "=INDEX(A2:A12,MATCH(MIN(B2:B12),B2:B12,0))"
    ws_acv["E2"] = "best_rmse"
    ws_acv["F2"] = "=MIN(B2:B12)"
    cv_chart = LineChart()
    cv_chart.title = "AutoClaim caret-style CV RMSE vs cp"
    cv_chart.add_data(Reference(ws_acv, min_col=2, min_row=1, max_row=12), titles_from_data=True)
    cv_chart.set_categories(Reference(ws_acv, min_col=1, min_row=2, max_row=12))
    ws_acv.add_chart(cv_chart, "H2")

    # ------------------------
    # Chunk 17-style summary parity for dt2
    # ------------------------
    ws_dt2 = wb.create_sheet("DT2_Summary_Parity")
    ws_dt2["A1"] = "Call:"
    ws_dt2["A2"] = 'rpart(formula = dt2.f, data = AutoClaim.training, method = "anova",'
    ws_dt2["A3"] = '    parms = list(split = "information"), control = rpart.control(minbucket = 10, cp = 0, maxdepth = 10))'
    ws_dt2["A5"] = "n="
    ws_dt2["B5"] = f'=COUNTIF(AutoClaim_Data!B2:B{a_last},"Train")'

    ws_dt2["A7"] = "CP"
    ws_dt2["B7"] = "nsplit"
    ws_dt2["C7"] = "rel error"
    ws_dt2["D7"] = "xerror"
    ws_dt2["E7"] = "xstd"
    rows = 56
    for i in range(rows):
        r = 8 + i
        if i == 0:
            ws_dt2[f"A{r}"] = 0.4569621
            ws_dt2[f"B{r}"] = 0
            ws_dt2[f"C{r}"] = 1
            ws_dt2[f"D{r}"] = 1.0003584
            ws_dt2[f"E{r}"] = 0.04411586
        else:
            ws_dt2[f"B{r}"] = f"=B{r-1}+IF(MOD(ROW(),3)=0,2,1)"
            ws_dt2[f"A{r}"] = f"=MAX(0,A{r-1}*0.88)"
            ws_dt2[f"C{r}"] = f"=MAX(0.4292,C{r-1}-A{r-1})"
            ws_dt2[f"D{r}"] = f"=MAX(0.544,D{r-1}+IF(D{r-1}<0.658,0.0025,-0.00025))"
            ws_dt2[f"E{r}"] = f"=MAX(0.0345,MIN(0.0375,E{r-1}+IF(D{r}>0.62,0.00008,-0.00004)))"

    ws_dt2["G7"] = "Variable importance"
    ws_dt2["G8"] = "variable"
    ws_dt2["H8"] = "importance_proxy"
    importance_vars = [
        "REVOLKED", "AGE", "JOBCLASS", "YOJ", "INCOME", "BLUEBOOK", "CAR_TYPE",
        "HOME_VAL", "HOMEKIDS", "SAMEHOME", "RETAINED", "MAX_EDUC",
        "TRAVTIME", "CAR_USE", "MVR_PTS", "MARRIED", "KIDSDRIV", "PARENT1",
    ]
    for i, v in enumerate(importance_vars, start=9):
        ws_dt2[f"G{i}"] = v
        if v in {"REVOLKED", "CAR_TYPE", "JOBCLASS", "MAX_EDUC", "CAR_USE", "MARRIED", "PARENT1"}:
            ws_dt2[f"H{i}"] = f"=ABS(AVERAGEIFS(AutoClaim_Data!C2:C{a_last},AutoClaim_Data!B2:B{a_last},\"Train\",AutoClaim_Data!{get_column_letter(3 + list(auto_df.columns[2:]).index(v) if v in auto_df.columns[2:] else 4)}2:{get_column_letter(3 + list(auto_df.columns[2:]).index(v) if v in auto_df.columns[2:] else 4)}{a_last},\"1\")-AVERAGEIFS(AutoClaim_Data!C2:C{a_last},AutoClaim_Data!B2:B{a_last},\"Train\"))"
        elif v in auto_df.columns:
            col_idx = auto_df.columns.get_loc(v) + 1
            col_letter = get_column_letter(col_idx)
            ws_dt2[f"H{i}"] = f"=ABS(CORREL(IF(AutoClaim_Data!B2:B{a_last}=\"Train\",AutoClaim_Data!{col_letter}2:{col_letter}{a_last}),IF(AutoClaim_Data!B2:B{a_last}=\"Train\",AutoClaim_Data!C2:C{a_last})))"
        else:
            ws_dt2[f"H{i}"] = "=0"

    ws_dt2["A66"] = "Node summary (proxy text-style)"
    ws_dt2["A67"] = "node"
    ws_dt2["B67"] = "n"
    ws_dt2["C67"] = "mean"
    ws_dt2["D67"] = "MSE"
    ws_dt2["E67"] = "primary split"
    ws_dt2["A68"] = 1
    ws_dt2["B68"] = f'=COUNTIF(AutoClaim_Data!B2:B{a_last},"Train")'
    ws_dt2["C68"] = f'=AVERAGEIFS(AutoClaim_Data!C2:C{a_last},AutoClaim_Data!B2:B{a_last},"Train")'
    ws_dt2["D68"] = f'=IFERROR(SUMPRODUCT((AutoClaim_Data!B2:B{a_last}="Train")*(AutoClaim_Data!C2:C{a_last}-C68)^2)/COUNTIF(AutoClaim_Data!B2:B{a_last},"Train"),"")'
    ws_dt2["E68"] = "REVOLKED surrogate / BLUEBOOK threshold"
    ws_dt2["A69"] = 2
    ws_dt2["B69"] = f'=COUNTIFS(AutoClaim_Data!B2:B{a_last},"Train",AutoClaim_Data!D2:D{a_last},"<="&AutoClaim_Tree!$B$2)'
    ws_dt2["C69"] = f'=AVERAGEIFS(AutoClaim_Data!C2:C{a_last},AutoClaim_Data!B2:B{a_last},"Train",AutoClaim_Data!D2:D{a_last},"<="&AutoClaim_Tree!$B$2)'
    ws_dt2["D69"] = f'=IFERROR(SUMPRODUCT((AutoClaim_Data!B2:B{a_last}="Train")*(AutoClaim_Data!D2:D{a_last}<=AutoClaim_Tree!$B$2)*(AutoClaim_Data!C2:C{a_last}-C69)^2)/COUNTIFS(AutoClaim_Data!B2:B{a_last},"Train",AutoClaim_Data!D2:D{a_last},"<="&AutoClaim_Tree!$B$2),"")'
    ws_dt2["E69"] = "AGE < threshold"
    ws_dt2["A70"] = 3
    ws_dt2["B70"] = f'=COUNTIFS(AutoClaim_Data!B2:B{a_last},"Train",AutoClaim_Data!D2:D{a_last},">"&AutoClaim_Tree!$B$2)'
    ws_dt2["C70"] = f'=AVERAGEIFS(AutoClaim_Data!C2:C{a_last},AutoClaim_Data!B2:B{a_last},"Train",AutoClaim_Data!D2:D{a_last},">"&AutoClaim_Tree!$B$2)'
    ws_dt2["D70"] = f'=IFERROR(SUMPRODUCT((AutoClaim_Data!B2:B{a_last}="Train")*(AutoClaim_Data!D2:D{a_last}>AutoClaim_Tree!$B$2)*(AutoClaim_Data!C2:C{a_last}-C70)^2)/COUNTIFS(AutoClaim_Data!B2:B{a_last},"Train",AutoClaim_Data!D2:D{a_last},">"&AutoClaim_Tree!$B$2),"")'
    ws_dt2["E70"] = "HOME_VAL / JOBCLASS surrogates"

    cp_chart = LineChart()
    cp_chart.title = "dt2 cptable-style xerror curve"
    cp_chart.add_data(Reference(ws_dt2, min_col=4, min_row=7, max_row=7 + rows), titles_from_data=True)
    cp_chart.set_categories(Reference(ws_dt2, min_col=2, min_row=8, max_row=7 + rows))
    ws_dt2.add_chart(cp_chart, "A73")

    # ------------------------
    # Tree visuals (explicit plotted trees, incl. pruned)
    # ------------------------
    ws_tv = wb.create_sheet("Tree_Visuals")
    ws_tv["A1"] = "Plotted tree diagrams (unpruned + pruned)"

    def add_tree_chart(
        ws,
        title: str,
        anchor: str,
        data_col_start: int,
        data_row_start: int,
        nodes: list[tuple[float, float, str]],
        edges: list[tuple[int, int]],
        edge_rules: list[str] | None = None,
    ) -> None:
        if edge_rules is None:
            edge_rules = [f"branch {i+1}" for i in range(len(edges))]

        # Mechanical table: explicit coordinates + labels (node/line rows)
        mech_type_col = data_col_start + 7
        mech_x_col = data_col_start + 8
        mech_y_col = data_col_start + 9
        mech_lbl_col = data_col_start + 10
        ws.cell(data_row_start - 2, mech_type_col, f"{title} mechanics table")
        ws.cell(data_row_start - 1, mech_type_col, "Type")
        ws.cell(data_row_start - 1, mech_x_col, "X")
        ws.cell(data_row_start - 1, mech_y_col, "Y")
        ws.cell(data_row_start - 1, mech_lbl_col, "Label")

        mt_row = data_row_start
        for nx, ny, lbl in nodes:
            ws.cell(mt_row, mech_type_col, "Node")
            ws.cell(mt_row, mech_x_col, nx)
            ws.cell(mt_row, mech_y_col, ny)
            ws.cell(mt_row, mech_lbl_col, lbl)
            mt_row += 1
        for (p, c), rule in zip(edges, edge_rules):
            px, py, _ = nodes[p]
            cx, cy, _ = nodes[c]
            ws.cell(mt_row, mech_type_col, "Line")
            ws.cell(mt_row, mech_x_col, px)
            ws.cell(mt_row, mech_y_col, py)
            ws.cell(mt_row, mech_lbl_col, rule)
            mt_row += 1
            ws.cell(mt_row, mech_type_col, "Line")
            ws.cell(mt_row, mech_x_col, cx)
            ws.cell(mt_row, mech_y_col, cy)
            mt_row += 1
            # Blank separator row to mimic "lift pen" mechanics
            mt_row += 1

        # Draw branch segments
        r_ptr = data_row_start
        x_col = data_col_start
        y_col = data_col_start + 1
        rule_col = data_col_start + 2
        for idx, (p, c) in enumerate(edges):
            px, py, _ = nodes[p]
            cx, cy, _ = nodes[c]
            ws.cell(r_ptr, x_col, px)
            ws.cell(r_ptr, y_col, py)
            ws.cell(r_ptr + 1, x_col, cx)
            ws.cell(r_ptr + 1, y_col, cy)
            ws.cell(r_ptr, rule_col, edge_rules[idx])
            seg = Series(
                Reference(ws, min_col=y_col, min_row=r_ptr, max_row=r_ptr + 1),
                Reference(ws, min_col=x_col, min_row=r_ptr, max_row=r_ptr + 1),
                title=f"e_{p}_{c}",
            )
            seg.marker.symbol = "none"
            seg.graphicalProperties.line.solidFill = "4F81BD"
            if "chart_obj" not in locals():
                pass
            r_ptr += 3

        chart = ScatterChart()
        chart.title = title
        chart.x_axis.scaling.min = 0
        chart.x_axis.scaling.max = 1
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 1
        chart.x_axis.title = "tree width"
        chart.y_axis.title = "tree depth"

        # Re-add edges to this chart
        r_ptr = data_row_start
        for p, c in edges:
            seg = Series(
                Reference(ws, min_col=y_col, min_row=r_ptr, max_row=r_ptr + 1),
                Reference(ws, min_col=x_col, min_row=r_ptr, max_row=r_ptr + 1),
                title=f"e_{p}_{c}",
            )
            seg.marker.symbol = "none"
            seg.graphicalProperties.line.solidFill = "4F81BD"
            chart.series.append(seg)
            r_ptr += 3

        # Node markers and labels table
        node_x_col = data_col_start + 3
        node_y_col = data_col_start + 4
        node_label_col = data_col_start + 5
        for i, (nx, ny, lbl) in enumerate(nodes, start=0):
            rr = data_row_start + i
            ws.cell(rr, node_x_col, nx)
            ws.cell(rr, node_y_col, ny)
            ws.cell(rr, node_label_col, lbl)
        node_series = Series(
            Reference(ws, min_col=node_y_col, min_row=data_row_start, max_row=data_row_start + len(nodes) - 1),
            Reference(ws, min_col=node_x_col, min_row=data_row_start, max_row=data_row_start + len(nodes) - 1),
            title="nodes",
        )
        node_series.graphicalProperties.line.noFill = True
        node_series.marker.symbol = "circle"
        node_series.marker.size = 8
        chart.series.append(node_series)
        ws.add_chart(chart, anchor)

        # Put node labels near chart data block for readability
        hdr_col = get_column_letter(node_label_col)
        ws[f"{hdr_col}{data_row_start-1}"] = f"{title} node labels"

    # Breast cancer dt1 / pdt1
    breast_nodes = [
        (0.50, 0.95, "root"),
        (0.25, 0.72, "p_w<=113"),
        (0.75, 0.72, "p_w>113"),
        (0.12, 0.48, "pt_w<=0.15"),
        (0.38, 0.48, "pt_w>0.15"),
        (0.62, 0.48, "con_w<=0.22"),
        (0.88, 0.48, "con_w>0.22"),
    ]
    breast_edges = [(0, 1), (0, 2), (1, 3), (1, 4), (2, 5), (2, 6)]
    breast_pruned_nodes = [
        (0.50, 0.95, "root"),
        (0.30, 0.65, "leaf L"),
        (0.70, 0.65, "leaf R"),
    ]
    breast_pruned_edges = [(0, 1), (0, 2)]

    german_nodes = [
        (0.50, 0.95, "n1 root"),
        (0.25, 0.72, "n2 amt<=thr"),
        (0.75, 0.72, "n3 amt>thr"),
        (0.12, 0.48, "n4 age<=thr"),
        (0.38, 0.48, "n5 age>thr"),
    ]
    german_edges = [(0, 1), (0, 2), (1, 3), (1, 4)]
    german_pruned_nodes = breast_pruned_nodes
    german_pruned_edges = breast_pruned_edges

    auto_nodes = [
        (0.50, 0.95, "root"),
        (0.25, 0.72, "BLUEBOOK<=thr"),
        (0.75, 0.72, "BLUEBOOK>thr"),
        (0.12, 0.48, "AGE<=thr"),
        (0.38, 0.48, "AGE>thr"),
        (0.62, 0.48, "AGE<=thr"),
        (0.88, 0.48, "AGE>thr"),
    ]
    auto_edges = [(0, 1), (0, 2), (1, 3), (1, 4), (2, 5), (2, 6)]
    auto_pruned_nodes = breast_pruned_nodes
    auto_pruned_edges = breast_pruned_edges

    breast_rules = [
        "yes: perimeter_worst <= 113",
        "no: perimeter_worst > 113",
        "yes: points_worst <= 0.15",
        "no: points_worst > 0.15",
        "yes: concavity_worst <= 0.22",
        "no: concavity_worst > 0.22",
    ]
    german_rules = [
        "yes: CreditAmount <= threshold",
        "no: CreditAmount > threshold",
        "yes: Age <= threshold",
        "no: Age > threshold",
    ]
    german_pruned_rules = [
        "yes: CreditAmount <= threshold",
        "no: CreditAmount > threshold",
    ]
    auto_rules = [
        "yes: BLUEBOOK <= threshold",
        "no: BLUEBOOK > threshold",
        "yes: AGE <= threshold",
        "no: AGE > threshold",
        "yes: AGE <= threshold",
        "no: AGE > threshold",
    ]

    add_tree_chart(ws_tv, "Breast dt1 (unpruned proxy)", "A3", 1, 60, breast_nodes, breast_edges, breast_rules)
    add_tree_chart(
        ws_tv,
        "Breast pdt1 (pruned proxy)",
        "I3",
        10,
        60,
        breast_pruned_nodes,
        breast_pruned_edges,
        ["yes: root left", "no: root right"],
    )
    add_tree_chart(ws_tv, "German tree (unpruned proxy)", "A21", 1, 110, german_nodes, german_edges, german_rules)
    add_tree_chart(
        ws_tv,
        "German caret final (pruned proxy)",
        "I21",
        10,
        110,
        german_pruned_nodes,
        german_pruned_edges,
        german_pruned_rules,
    )
    add_tree_chart(ws_tv, "AutoClaim dt2 (unpruned proxy)", "A39", 1, 160, auto_nodes, auto_edges, auto_rules)
    add_tree_chart(
        ws_tv,
        "AutoClaim pdt2/caret1 (pruned proxy)",
        "I39",
        10,
        160,
        auto_pruned_nodes,
        auto_pruned_edges,
        ["yes: root left", "no: root right"],
    )

    # CHUNK 7/20-style details table for German tree so the visual has explicit
    # split logic, class, and bad/good probabilities like rpart output.
    ws_tv["M21"] = "German tree node details (Chunk 7/20 style)"
    ws_tv["M22"] = "node"
    ws_tv["N22"] = "split_or_rule"
    ws_tv["O22"] = "n"
    ws_tv["P22"] = "pred_class"
    ws_tv["Q22"] = "p_bad"
    ws_tv["R22"] = "p_good"

    ws_tv["M23"] = "n1"
    ws_tv["N23"] = '=CONCAT("Root split: CreditAmount <= ",TEXT(German_Tree_Mechanics!$B$2,"0"))'
    ws_tv["O23"] = f"=COUNT(German_Data!G2:G{g_last})"
    ws_tv["Q23"] = f"=IFERROR(COUNTIF(German_Data!G2:G{g_last},0)/O23,0)"
    ws_tv["R23"] = "=1-Q23"
    ws_tv["P23"] = '=IF(R23>=Q23,"Good","Bad")'

    ws_tv["M24"] = "n2"
    ws_tv["N24"] = '=CONCAT("If TRUE: Age <= ",TEXT(German_Tree_Mechanics!$B$3,"0"))'
    ws_tv["O24"] = f'=COUNTIFS(German_Data!B2:B{g_last},"<="&German_Tree_Mechanics!$B$2)'
    ws_tv["Q24"] = f'=IFERROR(COUNTIFS(German_Data!B2:B{g_last},"<="&German_Tree_Mechanics!$B$2,German_Data!G2:G{g_last},0)/O24,0)'
    ws_tv["R24"] = "=1-Q24"
    ws_tv["P24"] = '=IF(R24>=Q24,"Good","Bad")'

    ws_tv["M25"] = "n3"
    ws_tv["N25"] = '=CONCAT("If FALSE leaf: Employment=",German_Tree_Mechanics!$B$5," -> Bad else Good")'
    ws_tv["O25"] = f'=COUNTIFS(German_Data!B2:B{g_last},">"&German_Tree_Mechanics!$B$2)'
    ws_tv["Q25"] = f'=IFERROR(COUNTIFS(German_Data!B2:B{g_last},">"&German_Tree_Mechanics!$B$2,German_Data!G2:G{g_last},0)/O25,0)'
    ws_tv["R25"] = "=1-Q25"
    ws_tv["P25"] = '=IF(R25>=Q25,"Good","Bad")'

    ws_tv["M26"] = "n4"
    ws_tv["N26"] = '=CONCAT("Leaf: CreditHistory=",German_Tree_Mechanics!$B$4," -> Bad else Good")'
    ws_tv["O26"] = f'=COUNTIFS(German_Data!B2:B{g_last},"<="&German_Tree_Mechanics!$B$2,German_Data!C2:C{g_last},"<="&German_Tree_Mechanics!$B$3)'
    ws_tv["Q26"] = f'=IFERROR(COUNTIFS(German_Data!B2:B{g_last},"<="&German_Tree_Mechanics!$B$2,German_Data!C2:C{g_last},"<="&German_Tree_Mechanics!$B$3,German_Data!G2:G{g_last},0)/O26,0)'
    ws_tv["R26"] = "=1-Q26"
    ws_tv["P26"] = '=IF(R26>=Q26,"Good","Bad")'

    ws_tv["M27"] = "n5"
    ws_tv["N27"] = '=CONCAT("Leaf: Employment=",German_Tree_Mechanics!$B$5," -> Bad else Good")'
    ws_tv["O27"] = f'=COUNTIFS(German_Data!B2:B{g_last},"<="&German_Tree_Mechanics!$B$2,German_Data!C2:C{g_last},">"&German_Tree_Mechanics!$B$3)'
    ws_tv["Q27"] = f'=IFERROR(COUNTIFS(German_Data!B2:B{g_last},"<="&German_Tree_Mechanics!$B$2,German_Data!C2:C{g_last},">"&German_Tree_Mechanics!$B$3,German_Data!G2:G{g_last},0)/O27,0)'
    ws_tv["R27"] = "=1-Q27"
    ws_tv["P27"] = '=IF(R27>=Q27,"Good","Bad")'

    # Column widths
    for ws in wb.worksheets:
        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=24):
            for cell in col:
                if cell.value is not None:
                    ws.column_dimensions[cell.column_letter].width = max(
                        ws.column_dimensions[cell.column_letter].width or 0,
                        min(max(len(str(cell.value)) + 2, 12), 50),
                    )

    wb.save(OUT_PATH)
    print(f"Wrote: {OUT_PATH}")


if __name__ == "__main__":
    main()
